
import os
import re
import sys
import time
import uuid
import json
import hashlib
import gc
from contextlib import contextmanager
from datetime import datetime, timezone
from typing import TypedDict, Literal, List, Dict, Any, Optional, Callable, Annotated, Tuple
from dataclasses import dataclass

from langgraph.graph import StateGraph, END
from langgraph.prebuilt import ToolNode
from langchain_core.messages import HumanMessage, SystemMessage, AIMessage, ToolMessage, AIMessageChunk

import chromadb
import openai
import requests

from config import settings
from utils.logger import get_logger
from utils.pdf_utils import HAS_PDF_TEXT, extract_pdf_text
logger = get_logger()
from services.inproc_openai_client import get_inproc_openai_client

from core.llm.builder import get_global_abort_manager, _GlobalAbortManager


from config.prompts import (
    CLASSIFY_PROMPT,
    SUMMARY_PROMPT,
    IMAGE_SUMMARY_PROMPT,
    IMAGE_OCR_PROMPT,
    MAP_PROMPT,
    REDUCE_PROMPT,
    INTENT_DETECTION_PROMPT,
    INTENT_DETECTION_SYSTEM_PROMPT,
    VIEW_DETAIL_INTENT_PROMPT,
    REWRITE_QUERY_PROMPT,
    AMBIGUOUS_QUERY_PROMPT,
    SUMMARIZE_SINGLE_FILE_PROMPT,
    SUMMARIZE_TOPICS_PROMPT,
    SUMMARIZE_ALL_PROMPT,
    CAPABILITY_QUERY_PROMPT,
    NO_RESULT_PROMPT,
    CHAT_FALLBACK_PROMPT,
    PLANNER_PROMPT,
    FINAL_ANSWER_PROMPT,
    EFFICIENT_ASSISTANT_PROMPT,
    get_prompt,
    normalize_prompt_language,
)
from core.retrieval.filename_canonicalizer import (
    compact_filename_key,
    has_plausible_filename_extension,
    is_descriptive_filename_phrase,
    is_reference_filename_placeholder,
    looks_like_thematic_lookup_candidate,
    looks_like_specific_filename_candidate,
    normalize_filename_candidate,
)
from core.retrieval.lookup_terms import extract_filelike_candidates

OCR_MAX_TOKENS_DEFAULT = int(os.getenv("OCR_MAX_TOKENS", "2200"))


AUDIO_EXTENSIONS = {'.mp3', '.wav', '.flac', '.aac', '.ogg', '.m4a', '.wma', '.aiff', '.ape', '.m4v', '.mp4'}

IMAGE_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp', '.svg', '.ico', '.tiff', '.heic'}

_INVENTORY_EXTENSION_GROUPS: Dict[str, Tuple[str, ...]] = {
    "pdf": (".pdf",), "pdfs": (".pdf",),
    "csv": (".csv",), "csvs": (".csv",),
    "tsv": (".tsv",), "tsvs": (".tsv",),
    "txt": (".txt",), "txts": (".txt",),
    "md": (".md",),
    "json": (".json",),
    "xml": (".xml",),
    "html": (".html",),
    "doc": (".doc",), "docx": (".docx",), "docxs": (".docx",),
    "xls": (".xls",), "xlsx": (".xlsx",),
    "ppt": (".ppt",), "pptx": (".pptx",),
    "png": (".png",), "pngs": (".png",),
    "jpg": (".jpg",), "jpgs": (".jpg",),
    "jpeg": (".jpeg",), "jpegs": (".jpeg",),
    "gif": (".gif",), "gifs": (".gif",),
    "webp": (".webp",), "webps": (".webp",),
    "bmp": (".bmp",), "bmps": (".bmp",),
    "wav": (".wav",), "wavs": (".wav",),
    "mp3": (".mp3",), "mp3s": (".mp3",),
    "m4a": (".m4a",), "m4as": (".m4a",),
    "mp4": (".mp4",), "mp4s": (".mp4",),
    "mov": (".mov",), "movs": (".mov",),
    # Format families should enumerate the common file extensions users expect.
    "excel": (".xlsx", ".xls", ".numbers"),
    "spreadsheet": (".xlsx", ".xls", ".csv", ".tsv", ".numbers"),
    "spreadsheets": (".xlsx", ".xls", ".csv", ".tsv", ".numbers"),
    "table": (".xlsx", ".xls", ".csv", ".tsv", ".numbers"),
    "tables": (".xlsx", ".xls", ".csv", ".tsv", ".numbers"),
    "worksheet": (".xlsx", ".xls", ".csv", ".tsv", ".numbers"),
    "worksheets": (".xlsx", ".xls", ".csv", ".tsv", ".numbers"),
    "word": (".docx", ".doc"),
    "powerpoint": (".pptx", ".ppt"),
    "presentation": (".pptx", ".ppt"),
    "presentations": (".pptx", ".ppt"),
    "slide": (".pptx", ".ppt"),
    "slides": (".pptx", ".ppt"),
}

_INVENTORY_CATEGORY_TERMS = {
    "file", "files", "document", "documents", "doc", "docs",
    "image", "images", "photo", "photos", "picture", "pictures",
    "screenshot", "screenshots",
    "video", "videos", "audio", "audios", "recording", "recordings",
    "clip", "clips", "movie", "movies",
    "spreadsheet", "spreadsheets", "worksheet", "worksheets", "table", "tables",
    "excel", "word", "powerpoint", "presentation", "presentations", "slide", "slides",
}

_INVENTORY_SCOPE_TERMS = {
    "find", "show", "list", "display", "browse", "which", "what", "about", "get",
    "give", "retrieve", "fetch", "all", "every", "my", "mine", "me", "the", "a", "an",
    "do", "i", "have",
}

# Canonical category storage uses English keys for indexing/search.
_CATEGORY_ALIASES_TO_EN: Dict[str, str] = {
    # resume
    "简历": "resume", "履历": "resume", "个人简历": "resume",
    "cv": "resume", "resume": "resume", "resumes": "resume",
    # report
    "报告": "report", "分析报告": "report", "研究报告": "report", "报告书": "report",
    "report": "report", "reports": "report",
    # contract
    "合同": "contract", "协议": "contract", "合约": "contract", "协议书": "contract",
    "contract": "contract", "contracts": "contract", "agreement": "contract", "agreements": "contract",
    # note
    "笔记": "note", "备忘": "note", "备忘录": "note",
    "note": "note", "notes": "note", "memo": "note", "memos": "note",
    # manual
    "手册": "manual", "教程": "manual", "指南": "manual", "说明书": "manual", "操作手册": "manual",
    "manual": "manual", "manuals": "manual", "guide": "manual", "guides": "manual", "tutorial": "manual", "tutorials": "manual",
    # paper
    "论文": "paper", "学术论文": "paper", "研究论文": "paper", "期刊": "paper",
    "paper": "paper", "papers": "paper", "article": "paper", "articles": "paper", "journal": "paper",
    # presentation
    "演示": "presentation", "演示文稿": "presentation", "幻灯片": "presentation", "课件": "presentation", "ppt": "presentation",
    "presentation": "presentation", "presentations": "presentation", "slides": "presentation", "deck": "presentation",
    # data
    "数据": "data", "数据集": "data", "表格": "data", "统计表": "data",
    "dataset": "data", "datasets": "data", "data": "data", "table": "data", "tables": "data",
    # email
    "邮件": "email", "电子邮件": "email", "邮箱": "email",
    "email": "email", "emails": "email", "mail": "email", "mails": "email",
    # image
    "照片": "image", "图片": "image", "图像": "image", "相片": "image", "截图": "image",
    "image": "image", "images": "image", "photo": "image", "photos": "image", "picture": "image", "pictures": "image",
    # audio/video
    "视频/音频": "audio/video", "影音": "audio/video", "视频": "audio/video", "音频": "audio/video", "录音": "audio/video",
    "audio": "audio/video", "video": "audio/video", "media": "audio/video", "podcast": "audio/video", "recording": "audio/video",
    # book
    "书籍": "book", "图书": "book", "电子书": "book",
    "book": "book", "books": "book", "ebook": "book", "ebooks": "book",
    # code
    "代码": "code", "源码": "code", "脚本": "code", "程序": "code",
    "code": "code", "codes": "code", "source code": "code", "script": "code", "scripts": "code",
    # invoice
    "账单": "invoice", "发票": "invoice", "收据": "invoice", "付款单": "invoice",
    "invoice": "invoice", "invoices": "invoice", "bill": "invoice", "bills": "invoice", "receipt": "invoice", "receipts": "invoice",
    # quotation
    "报价": "quotation", "报价单": "quotation", "报价表": "quotation", "报价书": "quotation", "报价方案": "quotation",
    "quote": "quotation", "quotes": "quotation", "quotation": "quotation", "quotations": "quotation", "proposal": "quotation", "proposals": "quotation",
    # document
    "文档": "document", "资料": "document", "文稿": "document",
    "document": "document", "documents": "document", "doc": "document", "docs": "document",
    "all": "all", "全部": "all", "所有": "all", "文件": "all", "文档全部": "all",
    "其他": "other", "未知": "other", "other": "other", "unknown": "other",
}

_CATEGORY_CANONICAL_TO_ZH: Dict[str, str] = {
    "resume": "简历",
    "report": "报告",
    "contract": "合同",
    "note": "笔记",
    "manual": "手册",
    "paper": "论文",
    "presentation": "演示",
    "data": "数据",
    "email": "邮件",
    "image": "图片",
    "audio/video": "音视频",
    "book": "书籍",
    "code": "代码",
    "invoice": "发票",
    "quotation": "报价单",
    "document": "文档",
    "all": "全部",
    "other": "其他",
}

_CATEGORY_KEYWORD_HINTS: List[Tuple[str, Tuple[str, ...]]] = [
    ("resume", ("简历", "履历", "resume", "cv")),
    ("report", ("报告", "分析报告", "report")),
    ("contract", ("合同", "协议", "合约", "agreement", "contract")),
    ("note", ("笔记", "备忘", "memo", "note")),
    ("manual", ("手册", "指南", "教程", "说明书", "manual", "guide", "tutorial")),
    ("paper", ("论文", "期刊", "paper", "article", "journal")),
    ("presentation", ("演示", "演示文稿", "幻灯", "ppt", "slides", "presentation", "deck")),
    ("data", ("数据", "表格", "dataset", "data", "table")),
    ("email", ("邮件", "邮箱", "email", "mail")),
    ("image", ("图片", "照片", "截图", "image", "photo", "picture")),
    ("audio/video", ("视频", "音频", "录音", "audio", "video", "media")),
    ("book", ("书籍", "图书", "电子书", "book", "ebook")),
    ("code", ("代码", "源码", "脚本", "code", "script", "source")),
    ("invoice", ("发票", "账单", "收据", "invoice", "bill", "receipt")),
    ("quotation", ("报价", "报价单", "报价书", "quotation", "quote", "proposal")),
    ("document", ("文档", "资料", "文件", "document", "doc")),
    ("all", ("全部", "所有", "all")),
    ("other", ("其他", "未知", "other", "unknown")),
]



def _is_generic_ext_inventory_query(q: str, candidate: str) -> bool:
    """Check if *candidate* looks like a generic inventory phrase rather than
    a specific filename (e.g. 'csv files', 'all pdf documents', 'my screenshots').

    Previously duplicated as an inline closure inside both
    _extract_explicit_file_reference and _extract_lexical_features.
    Both callers now delegate here.
    """
    import re as _re_gei
    ql = (q or "").lower()
    cand = (normalize_filename_candidate(str(candidate or "").strip()) or str(candidate or "").strip()).lower()
    if not cand:
        return False
    q_tokens = [t for t in _re_gei.findall(r"[a-z0-9]+", ql) if t]
    has_file_scope = bool(
        _re_gei.search(
            r"\b(file|files|filename|filenames|document|documents|doc|docs"
            r"|spreadsheet|spreadsheets|table|tables|sheet|sheets)\b",
            ql, _re_gei.IGNORECASE,
        )
        or any(tok in q for tok in ("\u6587\u4ef6", "\u6587\u6863", "\u8868\u683c",
                                    "\u6570\u636e\u8868", "\u7535\u5b50\u8868\u683c"))
    )
    all_tokens = [t for t in _re_gei.findall(r"[a-z0-9]+", f"{ql} {cand}") if t]
    recognized_terms = set(_INVENTORY_EXTENSION_GROUPS.keys()) | set(_INVENTORY_CATEGORY_TERMS)
    has_cjk_inventory_term = any(
        tok in q or tok in cand
        for tok in ("\u8868\u683c", "\u6570\u636e\u8868", "\u7535\u5b50\u8868\u683c", "\u5de5\u4f5c\u8868")
    )
    if not any(t in recognized_terms for t in all_tokens) and not has_cjk_inventory_term:
        return False
    has_inventory_verb = bool(
        _re_gei.search(r"\b(find|show|list|display|browse|which|what|get|give|retrieve|fetch)\b", ql)
        or any(tok in q for tok in ("\u627e", "\u67e5", "\u5217\u51fa", "\u770b\u770b", "\u7ed9\u6211\u770b", "\u67e5\u770b", "\u54ea\u4e9b", "\u6709\u4ec0\u4e48"))
    )
    has_all_scope = bool(
        _re_gei.search(r"\b(all|every)\b", ql)
        or any(tok in q for tok in ("\u5168\u90e8", "\u6240\u6709"))
    )
    has_possessive_scope = bool(
        _re_gei.search(r"\b(my|mine)\b", ql)
        or any(tok in q for tok in ("\u6211\u7684",))
    )
    has_inventory_question = bool(
        _re_gei.search(r"\bdo\s+i\s+have\b", ql)
        or any(tok in q for tok in ("\u6211\u6709", "\u6211\u8fd9\u91cc\u6709", "\u6211\u6709\u54ea\u4e9b"))
    )
    has_inventory_scope = bool(
        has_file_scope
        or ((has_inventory_verb or has_all_scope or has_possessive_scope or has_inventory_question) and any(t in recognized_terms for t in q_tokens))
    )
    if not has_inventory_scope:
        return False
    cand_base = os.path.basename(cand)
    cand_has_ext = has_plausible_filename_extension(cand_base)
    cand_stem = os.path.splitext(cand_base)[0] if cand_has_ext else cand_base
    cand_tokens = [t for t in _re_gei.findall(r"[a-z0-9]+", cand_stem) if t]
    if not cand_tokens:
        return bool(
            any(tok in q for tok in ("\u8868\u683c", "\u6570\u636e\u8868", "\u7535\u5b50\u8868\u683c", "\u5de5\u4f5c\u8868"))
            and has_inventory_scope
        )
    cand_core_tokens = [t for t in cand_tokens if t not in _INVENTORY_SCOPE_TERMS]
    if not cand_core_tokens:
        return False
    if cand_has_ext:
        return all(t in {"all", "every"} for t in cand_core_tokens)
    return all(t in recognized_terms for t in cand_core_tokens)


def _looks_like_prepositional_topic_fragment(candidate: str) -> bool:
    """Reject semantic fragments such as "of my dog" as filename candidates."""
    cand = normalize_filename_candidate(str(candidate or "").strip()) or str(candidate or "").strip()
    cand = cand.strip(" \"'“”‘’.,;:!?}>")
    if not cand:
        return False
    cand_base = os.path.basename(cand)
    if has_plausible_filename_extension(cand_base):
        return False
    if any(ch in cand_base for ch in ("_", "-", "/", "\\")) or any(ch.isdigit() for ch in cand_base):
        return False
    return bool(
        re.match(
            r"^(?:of|for|with|about|regarding|concerning|related\s+to|containing|contains|in|inside|under|from)\b",
            cand.lower(),
            re.IGNORECASE,
        )
    )


def _normalize_category_en(raw: str, default: str = "other") -> str:
    import re as _re_cat
    v = str(raw or "").strip().lower()
    if not v:
        return default
    direct = _CATEGORY_ALIASES_TO_EN.get(v)
    if direct:
        return direct
    for canonical, hints in _CATEGORY_KEYWORD_HINTS:
        for h in hints:
            if len(h) <= 3 and h.isascii():
                if _re_cat.search(r'(?<![a-z])' + _re_cat.escape(h) + r'(?![a-z])', v):
                    return canonical
            else:
                if h in v:
                    return canonical
    return v


def _localize_category_label(category_en: str, language: str) -> str:
    lang = str(language or "en").strip().lower()
    if lang.startswith("zh"):
        return _CATEGORY_CANONICAL_TO_ZH.get(category_en, category_en)
    return category_en


# ── English Category Inference ────────────────────────────────────────────────
# Rule-based safety net: when LLM intent pass doesn't emit a category for
# English queries, these trigger-word lists map query content to the correct
# file-type filter.  Ordered from most-specific to least-specific.
_EN_CATEGORY_TRIGGERS: List[Tuple[str, List[str]]] = [
    ("audio/video", [
        "audio", "sound", "recording", "music", "song", "track", "mp3",
        "wav", "flac", "aac", "m4a", "ogg", "wma", "aiff",
        "podcast", "voice memo", "noise", "melody", "beat", "rhythm",
        "video", "movie", "film", "clip", "footage", "reel", "mp4",
    ]),
    ("image", [
        "image", "photo", "picture", "screenshot", "diagram", "chart",
        "graphic", "png", "jpg", "jpeg", "gif", "bmp", "tiff", "heic",
        "illustration", "figure", "drawing", "sketch", "mockup", "thumbnail",
        "wiring diagram", "schematic", "blueprint",
    ]),
    # ‘document’ covers reports, analysis, resumes, invoices, strategy docs, papers
    ("document", [
        "resume", " cv ", "curriculum vitae", "invoice", "receipt", "bill",
        "contract", "agreement", "thesis", "presentation", "slides", "spreadsheet",
        # strategy / business docs — fix for “GTM”, “strategy”, “analysis” misclassified as manual
        "strategy", "gtm", "roadmap", "plan", "research", "findings", "insights",
        "analysis", "report", "paper",
    ]),
    # ‘manual’ only for explicit product/technical manuals — NEVER triggered by “manager” or “strategy”
    ("manual", [
        "product manual", "user guide", "data sheet", "datasheet",
        "installation guide", "technical specification", "spec sheet",
    ]),
]

# Words that look like triggers but should NOT infer category (anti-patterns)
_EN_CATEGORY_ANTITRIGGERS: set = {
    "manager", "management", "managed", "managing",  # "manager" sounds like "manual" to LLMs
}


def _infer_category_from_english_query(query: str) -> str:
    """
    Infer a file-type category from English query signal words.

    Returns the canonical category string (e.g. "audio/video", "image",
    "document") or an empty string if no signal is detected.

    Design notes:
    - Pure Python, zero LLM calls – executes in microseconds.
    - Used as a post-LLM safety net in _normalize_intent_to_internal_en.
    - Does NOT override a category already set by the LLM.
    - Deliberately conservative: only fires on clear, unambiguous signals.
    • Anti-trigger guard: words like 'manager' must not trigger 'manual'.
    """
    if not query:
        return ""
    q_lower = query.lower()
    q = f" {q_lower} "          # pad so " cv " matches "my cv"
    for category, triggers in _EN_CATEGORY_TRIGGERS:
        for trigger in triggers:
            # Use word-boundary-aware matching via space padding
            t = trigger if " " in trigger else f" {trigger} "
            if t in q:
                # Anti-trigger guard: don't fire if an anti-pattern overwrites
                # e.g. "manual" trigger blocked when "manager" present
                blocked = any(f" {anti} " in q for anti in _EN_CATEGORY_ANTITRIGGERS)
                if blocked and category == "manual":
                    continue
                return category
    return ""
# ─────────────────────────────────────────────────────────────────────────────


def _smart_truncate_summary(text: str, max_len: int = 200) -> str:
    text = text.strip()
    if len(text) <= max_len:
        return text
    cut = text[:max_len].rfind(" ")
    if cut < max_len * 0.4:
        for sep in (
            "。", "，", "、", "；", "：", "！", "？",
            "\u300d", "\uff09", "\u3011", "\u300b", "\u201d", "\u2019",
            ". ", ", ", "; ", ": ", "! ", "? ",
            ".", ",", ";", ":", "!", "?",
            ")", "]", "}", "—", "–", "-",
        ):
            alt = text[:max_len].rfind(sep)
            if alt > cut:
                cut = alt
    if cut < max_len * 0.25:
        cut = max_len
    truncated = text[:cut].rstrip(
        " .,;:!?~-\u2013\u2014\u00b7\u2026"
        "\uff0c\u3002\u3001\uff1b\uff1a\uff01\uff1f\uff5e\u00b7\u2026"
        "\u300c\u300d\u300e\u300f\uff08\uff09\u3010\u3011\u300a\u300b\u2329\u232a\u201c\u201d\u2018\u2019"
        "()[]{}\"\\'`"
        "\n\t\r"
    )
    return truncated + "..." if truncated != text.strip() else truncated


def _env_truthy(name: str) -> bool:
    v = (os.getenv(name) or "").strip().lower()
    return v in {"1", "true", "yes", "on", "y"}


def _indexing_cooperative_yield() -> None:
    try:
        sec = float(os.getenv("FILEAGENT_INDEX_YIELD_SEC", "0.02") or "0.02")
    except Exception:
        sec = 0.02
    if sec > 0:
        time.sleep(min(float(sec), 0.5))

BCE_QUERY_PREFIX = "Represent this sentence for searching relevant passages: "
BCE_DOC_PREFIX = "Represent this document for retrieval: "


from tools import get_all_tools, get_tool, IntentRegistry
from tools.document_tools import (
    count_documents, 
    count_documents_files,
    search_documents,
    summarize_topics,
    set_kb_instance,
    get_kb_instance,
)

# LlamaIndex is optional and disabled by default to match the bundled runtime.
# Set FILEAGENT_ENABLE_LLAMA_INDEX_FALLBACK=1 only for explicit experiments.
SimpleDirectoryReader = None
SentenceSplitter = None
HAS_LLAMA_INDEX = False
if _env_truthy("FILEAGENT_ENABLE_LLAMA_INDEX_FALLBACK"):
    try:
        from llama_index.core import SimpleDirectoryReader
        from llama_index.core.node_parser import SentenceSplitter
        HAS_LLAMA_INDEX = True
    except ImportError:
        logger.info("llama_index fallback enabled but not installed; using built-in parsers only")

class AgentState(TypedDict):
    question: str
    
    query_type: Literal["stats", "search", "chat"]
    category_filter: Optional[str]
    keyword: Optional[str]
    
    search_results: List[Dict[str, Any]]
    reranked_results: List[Dict[str, Any]]
    db_stats: Dict[str, Any]
    
    final_answer: str
    source_files: List[Dict[str, Any]]




# ── FileKnowledgeBase extracted → core/kb/knowledge_base.py ──────────────────
from core.kb.knowledge_base import (
    FileKnowledgeBase,
    _normalize_category_en,
    _localize_category_label,
    _infer_category_from_english_query,
    _env_truthy,
    _indexing_cooperative_yield,
    _CATEGORY_ALIASES_TO_EN,
    _CATEGORY_CANONICAL_TO_ZH,
    _CATEGORY_KEYWORD_HINTS,
    AUDIO_EXTENSIONS,
    IMAGE_EXTENSIONS,
    OCR_MAX_TOKENS_DEFAULT,
)

# Backward-compat alias
KnowledgeBase = FileKnowledgeBase






_NO_SYSTEM_ROLE_MODELS = {"gemma", "gemma2", "gemma3", "gemma-3"}


_model_system_role_cache: Dict[str, bool] = {}



# ── LLM utilities extracted → core/llm/ ──────────────────────────────────────
from core.llm.utils import (
    _model_supports_system_role,
    _estimate_tokens,
    _chunk_text,
    _chunk_text_by_newlines,
    _approx_tokens_from_text,
    _approx_tokens_from_messages,
    summarize_long_tool_result,
    convert_messages_for_gemma,
    build_messages_for_model,
    stream_replace_markdown_links,
)
from core.llm.builder import (
    get_llm,
    get_llm_with_tools,
    ToolAgentState,
)
from core.session import SessionMemory



class FileAgent:
    
    def __init__(self, db_path: str = settings.DB_PATH, llm_manager: Optional[Any] = None):
        logger.info(f"初始化 FileAgent (工具调用版)...")

        self._llm_manager = llm_manager
        
        self.kb = FileKnowledgeBase(db_path)
        
        set_kb_instance(self.kb)
        
        self._session_state = SessionMemory()
        self.max_history = 10
        self._category_counts_cache: Dict[str, int] = {}
        self._category_counts_cache_ts: float = 0.0
        self._category_counts_cache_total: int = -1
        
        self.log_file = os.path.join(os.path.dirname(db_path), "agent_log.txt")
        self._init_log()
        
        # self.graph = self._build_graph()
        
        logger.info(f"FileAgent 初始化完成")
    
    def set_abort_flag(self, session_id: Optional[str] = None):
        get_global_abort_manager().set(session_id)
        logger.info(f"设置中断标志: session={session_id or 'default'}")
    
    def clear_abort_flag(self, session_id: Optional[str] = None):
        get_global_abort_manager().clear(session_id)
        logger.info(f"清除中断标志: session={session_id or 'default'}")
    
    def is_aborted(self, session_id: Optional[str] = None) -> bool:
        return get_global_abort_manager().is_aborted(session_id)
    
    def _init_log(self):
        from datetime import datetime
        with open(self.log_file, 'a', encoding='utf-8') as f:
            f.write(f"\n{'='*60}\n")
            f.write(f"Session Start: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"{'='*60}\n")
    
    def _log(self, message: str, log_type: str = "INFO"):
        from datetime import datetime
        timestamp = datetime.now().strftime('%H:%M:%S')
        log_line = f"[{timestamp}] [{log_type}] {message}\n"
        
        try:
            with open(self.log_file, 'a', encoding='utf-8') as f:
                f.write(log_line)
        except Exception as e:
            logger.error(f"写日志失败: {e}")
    
    def clear_history(self):
        self._session_state.clear_all()
        self._log("对话历史已清空")

    def close(self) -> None:
        try:
            if getattr(self, "kb", None) is not None:
                self.kb.close()
        except Exception as e:
            logger.warning(f"FileAgent close failed: {e}")

    def _get_history_ref(self, session_id: Optional[str]) -> List[Dict]:
        return self._session_state.get_history_ref(session_id)

    def _get_last_search_results_ref(self, session_id: Optional[str]) -> List[Dict]:
        return self._session_state.get_last_search_results_ref(session_id)

    def _get_recent_search_result_sets(self, session_id: Optional[str], *, limit: int = 4) -> List[List[Dict]]:
        return self._session_state.get_recent_search_result_sets(session_id, limit=limit)

    def _set_last_search_results(self, session_id: Optional[str], results: List[Dict]) -> None:
        self._session_state.set_last_search_results(session_id, results)

    def _set_count_scope_context(self, session_id: Optional[str], context: Optional[Dict[str, Any]]) -> None:
        self._session_state.set_count_scope_context(session_id, context)

    def _get_count_scope_context(self, session_id: Optional[str]) -> Optional[Dict[str, Any]]:
        return self._session_state.get_count_scope_context(session_id)

    def _clear_count_scope_context(self, session_id: Optional[str], reason: str = "manual") -> None:
        self._session_state.clear_count_scope_context(session_id, reason=reason)

    def _build_count_scope_from_sources(self, sources: list) -> Dict[str, Any]:
        return self._session_state.build_count_scope_from_sources(sources)

    def _set_followup_hint(
        self,
        session_id: Optional[str],
        *,
        action: str,
        params: Optional[Dict[str, Any]] = None,
        ttl_turns: int = 2,
        uses: int = 1,
    ) -> None:
        self._session_state.set_followup_hint(
            session_id,
            action=action,
            params=params,
            ttl_turns=ttl_turns,
            uses=uses,
        )

    def _get_followup_hint(self, session_id: Optional[str]) -> Optional[Dict[str, Any]]:
        return self._session_state.get_followup_hint(session_id)

    def _consume_followup_hint(self, session_id: Optional[str]) -> None:
        self._session_state.consume_followup_hint(session_id)

    def _clear_followup_hint(self, session_id: Optional[str], reason: str = "manual") -> None:
        self._session_state.clear_followup_hint(session_id, reason=reason)

    def _log_followup_guard(
        self,
        *,
        stage: str,
        decision: str,
        reason: str,
        session_id: Optional[str],
        question: str,
        action_before: str = "",
        action_after: str = "",
        hint_action: str = "",
        brief_followup: Optional[bool] = None,
        last_results_count: int = 0,
    ) -> None:
        self._session_state.log_followup_guard(
            stage=stage,
            decision=decision,
            reason=reason,
            session_id=session_id,
            question=question,
            action_before=action_before,
            action_after=action_after,
            hint_action=hint_action,
            brief_followup=brief_followup,
            last_results_count=last_results_count,
        )

    def _sync_session_active_paths(self, session_id: Optional[str], active_paths: Optional[List[str]]) -> bool:
        return self._session_state.sync_active_paths(session_id, active_paths)

    def _clear_session_runtime_state(
        self,
        session_id: Optional[str],
        *,
        clear_history: bool = True,
        clear_language: bool = False,
        reason: str = "manual",
    ) -> None:
        self._session_state.clear_session_runtime_state(
            session_id,
            clear_history=clear_history,
            clear_language=clear_language,
            reason=reason,
        )

    def _is_brief_followup_query(self, question: str, *, prompt_language: Optional[str] = None) -> bool:
        q = (question or "").strip()
        if not q:
            return False
        q_compact = (
            q.replace("。", "").replace("，", "").replace("！", "").replace("？", "")
            .replace("!", "").replace("?", "").strip().lower()
        )
        lang = self._resolve_prompt_language(prompt_language, question=q)
        if lang == "zh":
            short_hits = {
                "需要", "好的", "好", "继续", "然后呢", "再来", "确认下", "确认一下", "帮我确认下",
                "核对下", "核对一下", "检查下", "检查一下", "详细说说", "详细一点",
            }
            followup_markers = ["这", "这些", "上面", "刚才", "前面", "那个", "它", "帮我", "给我", "一下", "下", "继续", "再"]
            return (q_compact in short_hits) or (len(q) <= 24 and any(m in q for m in followup_markers))
        words = [w for w in q_compact.split() if w]
        short_hits_en = {
            "yes", "ok", "okay", "sure", "go on", "continue", "confirm", "verify",
            "check this", "help me confirm", "please confirm",
        }
        followup_markers_en = ["this", "these", "that", "them", "it", "previous", "above", "earlier", "help me", "please"]
        return (q_compact in short_hits_en) or (len(words) <= 10 and any(m in q_compact for m in followup_markers_en))

    def _looks_like_meta_followup_on_last_results(self, question: str, prompt_language: Optional[str] = None) -> bool:
        from core.intent_analyzer import IntentAnalyzer
        return IntentAnalyzer.looks_like_meta_followup_on_last_results(question, prompt_language)

    def _prompt(self, prompt_name: str, prompt_language: Optional[str]) -> str:
        return get_prompt(prompt_name, language=prompt_language)

    def _resolve_prompt_language(
        self,
        prompt_language: Optional[str],
        *,
        question: str = "",
        session_id: Optional[str] = None,
    ) -> str:
        def _normalize_or_empty(value: Optional[str]) -> str:
            raw = str(value or "").strip().lower()
            if raw.startswith("zh"):
                return "zh"
            if raw.startswith("en"):
                return "en"
            return ""

        def _detect_question_language(q: str) -> str:
            txt = (q or "").strip()
            if not txt:
                return ""
            
            has_cjk = any("\u4e00" <= ch <= "\u9fff" for ch in txt)
            if has_cjk:
                return "zh"
            
            has_latin = any(("a" <= ch <= "z") or ("A" <= ch <= "Z") for ch in txt)
            if has_latin and not has_cjk:
                return "en"
                
            return ""

        lang = _normalize_or_empty(prompt_language)
        if lang in {"zh", "en"}:
            return lang

        # Priority (fallback behavior):
        # 1) detect from current user question
        # 2) session remembered language
        # 3) ambiguous default -> English
        q_lang = _detect_question_language(question)
        if q_lang in {"zh", "en"}:
            return q_lang

        sid = (session_id or "").strip()
        if sid:
            cached = _normalize_or_empty(self._session_state.get_remembered_language(sid))
            if cached in {"zh", "en"}:
                return cached

        return "en"

    def _remember_prompt_language(self, session_id: Optional[str], prompt_language: str) -> None:
        self._session_state.remember_prompt_language(session_id, prompt_language)

    def _augment_query_for_retrieval(
        self,
        query: str,
        *,
        prompt_language: Optional[str],
        session_id: Optional[str] = None,
    ) -> str:
        """Delegate to core.retrieval.query_augmenter."""
        from core.retrieval.query_augmenter import augment_query_for_retrieval
        llm = self._get_llm_service(detailed=False, session_id=session_id, prompt_language="en")
        return augment_query_for_retrieval(
            query,
            llm_service=llm,
            last_results_fn=lambda: self._get_last_search_results_ref(session_id),
            session_id=session_id,
        )

    def _resolve_personal_attribute_query(
        self,
        query: str,
        *,
        session_id: Optional[str] = None,
        prompt_language: Optional[str] = None,
    ) -> Optional[Dict[str, str]]:
        import re as _re

        # ── Fast guard: must look like an attribute lookup ────────────────────
        _PRONOUN_PAT = _re.compile(
            r'\b(his|her|their|he|she|they|him)\b|他的?|她的?|他们的?', _re.IGNORECASE
        )
        _ATTR_PAT = _re.compile(
            r'\b(email|e-mail|phone|mobile|address|contact|tel|salary|birthday|wechat|linkedin|'
            r'home\s+address|residence|residential\s+address|'
            r'school|university|college|alma\s+mater|education|degree|major|graduat(?:e|ed|ion)|'
            r'employer|company|job|title|role|position)\b|'
            r'\b(his|her|their|its)\s+(home|residence|residential\s+address)\b|'
            r'\b(?:where\s+(?:does|do|is|are)|where).{0,24}\b(?:he|she|they|his|her|their).{0,24}'
            r'(?:live|lives|reside|resides|home|address|location)\b|'
            r'(邮箱|手机|电话|地址|联系方式|联系电话|生日|工资|薪资|邮件|微信|'
            r'家庭住址|住址|居住地|居住地址|住所|住宅|住哪|住在哪里|家在哪|家在哪里|'
            r'毕业院校|毕业学校|哪个学校|什么学校|毕业于|学历|学位|专业|公司|单位|雇主|职位|职务|岗位)',
            _re.IGNORECASE,
        )
        if not _ATTR_PAT.search(query):
            return None

        def _attr_from_query() -> str:
            am = _ATTR_PAT.search(query)
            return am.group(0).strip() if am else "contact info"

        def _canonicalize_attribute(value: str) -> str:
            lowered = str(value or "").strip().lower()
            if not lowered:
                return "contact info"
            if any(tok in lowered for tok in ("email", "e-mail", "邮箱", "邮件")):
                return "email"
            if any(tok in lowered for tok in ("phone", "mobile", "tel", "telephone", "电话", "手机", "手机号", "联系电话")):
                return "phone"
            if any(
                tok in lowered
                for tok in (
                    "address",
                    "location",
                    "home",
                    "residence",
                    "residential",
                    "live",
                    "reside",
                    "地址",
                    "家庭住址",
                    "住址",
                    "居住",
                    "住所",
                    "住宅",
                    "住哪",
                    "家在哪",
                    "位置",
                    "所在地",
                )
            ):
                return "address"
            if any(tok in lowered for tok in ("wechat", "微信")):
                return "wechat"
            if "linkedin" in lowered:
                return "linkedin"
            if any(tok in lowered for tok in ("salary", "工资", "薪资", "薪水")):
                return "salary"
            if any(tok in lowered for tok in ("birthday", "birth", "生日")):
                return "birthday"
            if any(tok in lowered for tok in ("school", "university", "college", "alma", "graduat", "毕业", "学校", "大学", "高校")):
                return "school"
            if any(tok in lowered for tok in ("education", "degree", "学历", "学位")):
                return "education"
            if any(tok in lowered for tok in ("major", "专业")):
                return "major"
            if any(tok in lowered for tok in ("employer", "company", "公司", "单位", "雇主")):
                return "company"
            if any(tok in lowered for tok in ("job", "title", "role", "position", "职位", "职务", "岗位")):
                return "position"
            return lowered

        def _query_contains_attribute(value: str, attribute: str) -> bool:
            text = str(value or "").strip().lower()
            attr = str(attribute or "").strip().lower()
            if not text or not attr:
                return False
            attr_terms = {
                "email": ("email", "e-mail", "邮箱", "邮件"),
                "phone": ("phone", "mobile", "tel", "telephone", "电话", "手机", "手机号"),
                "address": (
                    "address",
                    "location",
                    "home",
                    "residence",
                    "residential",
                    "住址",
                    "地址",
                    "居住",
                    "住所",
                    "住宅",
                    "家在",
                    "所在地",
                ),
                "wechat": ("wechat", "weixin", "微信"),
                "linkedin": ("linkedin",),
                "salary": ("salary", "工资", "薪资", "薪水"),
                "birthday": ("birthday", "birth", "生日"),
                "school": ("school", "university", "college", "alma", "graduat", "毕业", "学校", "大学"),
                "education": ("education", "degree", "学历", "学位"),
                "major": ("major", "专业"),
                "company": ("employer", "company", "公司", "单位", "雇主"),
                "position": ("job", "title", "role", "position", "职位", "职务", "岗位"),
            }
            terms = attr_terms.get(attr) or (attr,)
            return any(term and term in text for term in terms)

        def _force_attribute_into_query(value: str, entity: str, attribute: str) -> str:
            resolved = str(value or "").strip()
            attr = _canonicalize_attribute(attribute)
            if not attr or attr == "contact info" or _query_contains_attribute(resolved, attr):
                return resolved
            ent = str(entity or "").strip()
            if ent:
                return f"{ent} {attr}".strip()
            return f"{resolved} {attr}".strip()

        # ── Build context from last_results (file names + doc summaries) ──────
        last = self._get_last_search_results_ref(session_id) or []
        hist_ref = self._get_history_ref(session_id) or []
        recent_turns = list(hist_ref[-3:-1]) if hist_ref else []
        recent_user_queries = [
            str(item.get("q") or "").strip()
            for item in recent_turns
            if str(item.get("q") or "").strip()
        ]
        recent_answer_previews = [
            str(item.get("a") or "").strip()[:180]
            for item in recent_turns
            if str(item.get("a") or "").strip()
        ]

        ctx_lines = []
        for r in last[:5]:
            fn = str(r.get("file_name") or "").strip()
            ds = str(r.get("doc_summary") or "").strip()[:120]
            if fn:
                ctx_lines.append(f"- {fn}: {ds}" if ds else f"- {fn}")
        context_str = "\n".join(ctx_lines)
        recent_query_str = "\n".join(f"- {q}" for q in recent_user_queries[:3])
        recent_answer_str = "\n".join(f"- {a}" for a in recent_answer_previews[:2] if a)

        # ── No context: still let LLM extract explicit entity/attribute from the query ──
        # This helps first-turn queries such as "find PersonName's phone number"
        # without turning the path into a large rule soup.
        only_query_mode = not context_str and not recent_query_str

        # ── LLM call: entity extraction only, minimal prompt ─────────────────
        try:
            llm = self._get_llm_service(detailed=False, session_id=session_id, prompt_language="en")
            prompt_sections = []
            if recent_query_str:
                prompt_sections.append("Recent user queries:\n" + recent_query_str)
            if context_str:
                prompt_sections.append("Previous search returned these files:\n" + context_str)
            if recent_answer_str:
                prompt_sections.append("Recent assistant answer previews:\n" + recent_answer_str)
            prompt = (
                "\n\n".join(prompt_sections)
                + f'\n\nUser now asks: "{query}"\n\n'
                + "Task: identify the target person (or contact subject) and the attribute they want.\n"
                + "If the user uses a pronoun (his/her/he/she/他/她), resolve it from the recent conversation.\n"
                + "If the user explicitly names a person in the current query, preserve that person as the target.\n"
                + "Prefer the person explicitly mentioned in the recent user queries when available.\n"
                + "ATTRIBUTE must be one short canonical field such as: email, phone, address, location, home address, residence, wechat, linkedin, salary, birthday, school, education, major, company, position.\n"
                + "If the query is a first-turn query with no context, still extract the explicit person name if present.\n"
                + "Prefer the most canonical full name form. If the person is likely Chinese and you can confidently infer the Chinese full name from context or common transliteration, prefer that full name.\n"
                + "Reply in this exact format (one line, no explanation):\n"
                + "ENTITY: <person name or unknown> | ATTRIBUTE: <attribute or unknown> | RESOLVED_QUERY: <best concise search query>\n"
                + "Examples:\n"
                + "ENTITY: PersonName | ATTRIBUTE: email | RESOLVED_QUERY: PersonName email\n"
                + "ENTITY: PersonName | ATTRIBUTE: phone | RESOLVED_QUERY: PersonName phone\n"
                + "ENTITY: PersonName | ATTRIBUTE: address | RESOLVED_QUERY: PersonName home address\n"
                + "ENTITY: PersonName | ATTRIBUTE: school | RESOLVED_QUERY: PersonName graduation school\n"
                + "If you cannot determine the person, set ENTITY to unknown and keep RESOLVED_QUERY close to the original query."
            )
            raw = (llm.generate(prompt) or "").strip()

            # Parse "ENTITY: PersonName | ATTRIBUTE: email | RESOLVED_QUERY: PersonName email"
            m = _re.search(
                r'ENTITY:\s*(.+?)\s*\|\s*ATTRIBUTE:\s*(.+?)\s*\|\s*RESOLVED_QUERY:\s*(.+)',
                raw, _re.IGNORECASE
            )
            if not m:
                # LLM returned unexpected format: redirect with original query
                return {"resolved_query": query, "entity": "", "attribute": _attr_from_query()}

            entity = m.group(1).strip()
            attribute = _canonicalize_attribute(m.group(2).strip())
            resolved_query = str(m.group(3) or "").strip()
            if not entity or entity.lower() in {"unknown", "n/a", "none", ""}:
                fallback_query = resolved_query if resolved_query and not only_query_mode else query
                return {
                    "resolved_query": fallback_query,
                    "entity": "",
                    "attribute": _canonicalize_attribute(_attr_from_query()),
                }

            if not resolved_query:
                resolved_query = f"{entity} {attribute}".strip()
            resolved_query = _force_attribute_into_query(resolved_query, entity, attribute)
            return {
                "resolved_query": resolved_query,
                "entity": entity,
                "attribute": attribute,
            }
        except Exception as _e:
            logger.warning(f"[resolve_pronoun] LLM failed: {_e}")
            # On LLM failure: still redirect to global search with original query
            return {"resolved_query": query, "entity": "", "attribute": _canonicalize_attribute(_attr_from_query())}


    def _resolve_pronoun_query(
        self,
        query: str,
        *,
        session_id: Optional[str] = None,
        prompt_language: Optional[str] = None,
    ) -> Optional[Dict[str, str]]:
        """
        Backward-compatible wrapper.
        Historically this only handled pronoun-based follow-ups, but the
        retrieval path now benefits from the same lightweight entity extraction
        for first-turn personal-attribute queries as well.
        """
        return self._resolve_personal_attribute_query(
            query,
            session_id=session_id,
            prompt_language=prompt_language,
        )


    @staticmethod
    def _strip_meta_for_rerank(query: str) -> str:
        """Delegate to core.retrieval.query_augmenter."""
        from core.retrieval.query_augmenter import strip_meta_for_rerank
        return strip_meta_for_rerank(query)


    def _blend_retrieval_query_with_original_cjk(self, retrieval_query: str, original_question: str) -> str:
        """Delegate to core.retrieval.query_augmenter."""
        from core.retrieval.query_augmenter import blend_retrieval_query_with_original_cjk
        return blend_retrieval_query_with_original_cjk(retrieval_query, original_question)


    def _looks_like_content_followup_on_prior_results(self, question: str) -> bool:
        from core.intent_analyzer import IntentAnalyzer
        return IntentAnalyzer.looks_like_content_followup_on_prior_results(question)

    def _anchor_retrieval_query_with_last_search(
        self, retrieval_query: str, original_question: str, session_id: Optional[str]
    ) -> str:
        """Delegate to core.retrieval.query_augmenter."""
        from core.retrieval.query_augmenter import anchor_retrieval_query_with_last_search
        try:
            last = self._get_last_search_results_ref(session_id) or []
        except Exception:
            last = []
        return anchor_retrieval_query_with_last_search(retrieval_query, original_question, last)


    def _normalize_intent_to_internal_en(
        self,
        question: str,
        result: dict,
        *,
        session_id: Optional[str] = None,
    ) -> dict:
        """
        Normalize legacy/raw intent payloads, then run the centralized validator
        so every call site gets the same correction layer.
        """
        from core.intent.normalizer import QueryNormalizer
        from core.intent.validator import IntentValidator

        normalized = QueryNormalizer.normalize(
            question,
            result,
            normalize_category_fn=self._normalize_category_name,
        )
        hist_ref = self._get_history_ref(session_id)
        active_paths = None
        if session_id:
            try:
                active_paths = list(self._session_state.get_state(session_id).active_paths or [])
            except Exception:
                active_paths = None

        return IntentValidator.validate(
            question,
            normalized,
            last_results=self._get_last_search_results_ref(session_id),
            history=list(hist_ref),
            active_paths=active_paths,
            prompt_language=self._resolve_prompt_language(None, question=question, session_id=session_id),
        )


    def _localize_text(
        self,
        text: str,
        *,
        target_language: str,
        session_id: Optional[str] = None,
    ) -> str:
        _ = target_language
        _ = session_id
        return str(text or "")

    def _post_localize_text(
        self,
        text: str,
        *,
        prompt_language: Optional[str],
        session_id: Optional[str] = None,
    ) -> str:
        _ = prompt_language
        _ = session_id
        return str(text or "")

    def _normalize_category_name(self, category: str) -> str:
        return _normalize_category_en(category, default="other")

    def _paper_category_likely_wrong_for_query(self, user_question: str, normalized_category: str) -> bool:
        """Delegate to core.retrieval.category_engine."""
        from core.retrieval.category_engine import paper_category_likely_wrong_for_query
        return paper_category_likely_wrong_for_query(user_question, normalized_category)


    def _report_category_likely_wrong_for_query(self, user_question: str, normalized_category: str) -> bool:
        """Delegate to core.retrieval.category_engine."""
        from core.retrieval.category_engine import report_category_likely_wrong_for_query
        return report_category_likely_wrong_for_query(user_question, normalized_category)


    def _collect_category_counts(self) -> Dict[str, int]:
        try:
            kb = get_kb_instance()
            from core.retrieval.category_engine import normalize_stored_category_name
            now = time.time()
            try:
                _ttl = float(os.getenv("FILEAGENT_CATEGORY_CACHE_TTL_SEC", "60") or "60")
            except Exception:
                _ttl = 60.0
            _ttl = max(5.0, min(_ttl, 3600.0))

            # ── Fast path: return cached result if still within TTL ──
            # collection.count() is also skipped to avoid an extra Chroma call
            # on every query. TTL invalidation (60 s default) is sufficient for
            # intent-analysis stats that don't need real-time accuracy.
            if (
                self._category_counts_cache
                and (now - float(self._category_counts_cache_ts or 0.0)) < _ttl
            ):
                return dict(self._category_counts_cache)

            # Cache expired — rebuild (also validate against current collection size)
            total_count = int(kb.collection.count())
            if (
                self._category_counts_cache
                and self._category_counts_cache_total == total_count
            ):
                # Collection size unchanged: refresh timestamp, skip full rebuild
                self._category_counts_cache_ts = now
                return dict(self._category_counts_cache)

            raw = kb.count_all_categories() or {}
            merged: Dict[str, int] = {}
            for cat, cnt in raw.items():
                k = normalize_stored_category_name(str(cat or "other"))
                merged[k] = merged.get(k, 0) + int(cnt or 0)
            self._category_counts_cache = dict(merged)
            self._category_counts_cache_total = total_count
            self._category_counts_cache_ts = now
            return merged
        except Exception:
            if self._category_counts_cache:
                return dict(self._category_counts_cache)
            return {}


    def _get_rule_category_keywords(self) -> List[str]:
        base = list(_CATEGORY_ALIASES_TO_EN.keys())
        dynamic = sorted(
            self._collect_category_counts().items(),
            key=lambda x: (-int(x[1]), len(str(x[0]))),
        )
        kws: List[str] = []
        for cat in base + [str(k) for k, _ in dynamic]:
            raw = str(cat or "").strip()
            norm = self._normalize_category_name(raw)
            if raw and raw not in kws:
                kws.append(raw)
            if norm and norm not in kws:
                kws.append(norm)

        kws.sort(key=lambda x: len(x), reverse=True)
        return kws

    def _is_previous_results_reference(self, question: str, prompt_language: Optional[str] = None) -> bool:
        from core.intent_analyzer import IntentAnalyzer, IntentKeywords
        return IntentAnalyzer._is_kw_match(question.lower(), IntentKeywords.PREV_REF_KWS, 'all')

    def _is_all_files_list_query(self, question: str, prompt_language: Optional[str] = None) -> bool:
        from core.intent_analyzer import IntentAnalyzer
        return IntentAnalyzer._looks_like_all_files_list_query(question)

    def _looks_like_scoped_file_search_query(self, question: str, prompt_language: Optional[str] = None) -> bool:
        from core.intent_analyzer import IntentAnalyzer
        return IntentAnalyzer._looks_like_scoped_file_search_query(question)

    def _looks_like_generic_inventory_query(self, question: str) -> bool:
        return _is_generic_ext_inventory_query(question, question)

    def _is_generic_file_scope_category(self, category: str) -> bool:
        """Delegate to core.retrieval.category_engine."""
        from core.retrieval.category_engine import is_generic_file_scope_category
        return is_generic_file_scope_category(category)

    def _looks_like_folder_listing_query(self, question: str) -> bool:
        import re

        q = str(question or "").strip()
        if not q:
            return False
        ql = q.lower()
        has_folder_signal = bool(
            re.search(r"\b(folder|directory|dir)\b", ql)
            or any(tok in q for tok in ("目录", "文件夹"))
        )
        if not has_folder_signal:
            return False
        has_container_phrase = bool(
            re.search(
                r"\b(?:in|inside|under|within|from)\s+(?:the\s+)?(?:folder|directory|dir)\b"
                r"|\b(?:folder|directory|dir)\b\s+[^\n\r]{0,80}\b(?:with|containing)\b",
                ql,
                re.IGNORECASE,
            )
            or any(tok in q for tok in ("目录里", "目录中的", "目录下", "文件夹里", "文件夹中的", "文件夹下", "里面的", "下的"))
        )
        has_listing_target = bool(
            re.search(
                r"\b(files?|documents?|docs?|images?|photos?|pictures?|audio|recordings?|videos?|clips?|pdfs?|csvs?|xlsx|xls|wav|mp3|mp4)\b",
                ql,
                re.IGNORECASE,
            )
            or any(tok in q for tok in ("文件", "文档", "图片", "照片", "表格", "音频", "录音", "视频"))
        )
        return has_container_phrase and has_listing_target

    def _extract_explicit_file_reference(self, question: str) -> Optional[Dict[str, str]]:
        import re

        q = str(question or "").strip()
        if not q:
            return None
        if self._looks_like_folder_listing_query(q):
            return None

        def _looks_like_generic_extension_inventory_request(candidate: str = "") -> bool:
            # Delegates to module-level helper (deduplicated from _extract_explicit_file_reference)
            return _is_generic_ext_inventory_query(q, candidate)

        def _looks_like_folder_contents_query(candidate: str = "") -> bool:
            ql = q.lower()
            cand = str(candidate or "").strip().lower()
            combined = f"{ql} {cand}".strip()
            has_folder_signal = bool(
                re.search(r"\b(folder|directory|dir)\b", ql)
                or any(tok in q for tok in ("目录", "文件夹"))
            )
            if not has_folder_signal:
                return False
            has_content_scope = bool(
                re.search(
                    r"\b(files?|documents?|docs?|images?|photos?|pictures?|audio|recordings?|videos?|clips?|pdfs?|csvs?|xlsx|xls|wav|mp3|mp4)\b",
                    ql,
                    re.IGNORECASE,
                )
                or any(tok in q for tok in ("文件", "文档", "图片", "照片", "表格", "音频", "录音", "视频"))
            )
            has_container_phrase = bool(
                re.search(
                    r"\b(?:in|inside|under|within|from)\s+(?:the\s+)?(?:folder|directory|dir)\b"
                    r"|\b(?:folder|directory|dir)\b\s+(?:named\s+)?[^\n\r]{0,80}\b(?:with|containing)\b",
                    ql,
                    re.IGNORECASE,
                )
                or any(tok in combined for tok in ("目录里", "目录中的", "目录下", "文件夹里", "文件夹中的", "文件夹下", "里的", "里面的", "下的"))
            )
            return has_content_scope and has_container_phrase

        patterns = [
            r'[:：]\s*([A-Za-z0-9_][A-Za-z0-9_\-(). ]{2,}\.[A-Za-z0-9]{2,6})',
            r'["“]?([A-Za-z0-9_][A-Za-z0-9_\-(). ]{2,}\.[A-Za-z0-9]{2,6})["”]?',
        ]
        raw_name = ""
        extracted_from_filename_surface = False
        filename_surfaces = extract_filelike_candidates(q, max_candidates=8)
        if len(filename_surfaces) >= 2:
            raw_name = os.path.basename(normalize_filename_candidate(str(filename_surfaces[0] or "").strip()))
            extracted_from_filename_surface = bool(raw_name)
        for pat in patterns:
            if raw_name:
                break
            m = re.search(pat, q)
            if m:
                candidate = normalize_filename_candidate(str(m.group(1) or "").strip())
                expanded = extract_filelike_candidates(candidate, max_candidates=4)
                if expanded:
                    candidate = normalize_filename_candidate(str(expanded[0] or "").strip())
                if candidate and has_plausible_filename_extension(candidate):
                    raw_name = candidate
                    extracted_from_filename_surface = True
                    break

        if not raw_name:
            quote_guess = re.search(r'["“]([^\n\r"”]{2,80})["”]', q)
            if quote_guess:
                raw_name = normalize_filename_candidate(str(quote_guess.group(1) or "").strip())
                extracted_from_filename_surface = bool(raw_name)

        if not raw_name:
            colon_guess = re.search(r'[:：]\s*([^\n\r,，。!?？]{2,80})', q)
            if colon_guess:
                candidate = normalize_filename_candidate(str(colon_guess.group(1) or "").strip())
                if candidate and any(ch in candidate for ch in ["_", "-", ".", " "]):
                    raw_name = candidate
                    extracted_from_filename_surface = True
        if not raw_name:
            verb_tail_guess = re.search(
                r'^\s*(?:please\s+)?(?:help\s+me\s+(?:to\s+)?|can\s+you\s+)?'
                r'(?:find|show|search\s+for|look\s+for|get|open|locate)\s+'
                r'(?P<target>[A-Za-z0-9_][A-Za-z0-9_\-./]{1,79})\s*$',
                q,
                re.IGNORECASE,
            )
            if verb_tail_guess:
                candidate = normalize_filename_candidate(str(verb_tail_guess.group("target") or "").strip())
                if candidate and (
                    has_plausible_filename_extension(candidate)
                    or any(ch in candidate for ch in ("_", "-", "/", "\\"))
                    or any(ch.isdigit() for ch in candidate)
                    or looks_like_specific_filename_candidate(candidate)
                ):
                    raw_name = os.path.basename(candidate)
                    extracted_from_filename_surface = True
        if not raw_name:
            filename_candidates = filename_surfaces or extract_filelike_candidates(q, max_candidates=4)
            if filename_candidates:
                candidate = normalize_filename_candidate(str(filename_candidates[0] or "").strip())
                lower = candidate.lower()
                if (
                    candidate
                    and not lower.startswith(("http://", "https://", "www."))
                    and "@" not in candidate
                    and not _looks_like_folder_contents_query(candidate)
                    and not looks_like_thematic_lookup_candidate(q, candidate)
                    and (
                        has_plausible_filename_extension(candidate)
                        or looks_like_specific_filename_candidate(candidate)
                    )
                ):
                    raw_name = os.path.basename(candidate)
                    extracted_from_filename_surface = True
        if not raw_name:
            return None

        raw_name = raw_name.strip(" \"'“”‘’.,;:!?}>")
        if _looks_like_prepositional_topic_fragment(raw_name):
            return None
        folder_prefix_match = re.match(
            r'^(?:folder|directory|dir)\s+(?:named\s+|called\s+)?(.+)$',
            raw_name,
            re.IGNORECASE,
        )
        if folder_prefix_match:
            raw_name = str(folder_prefix_match.group(1) or "").strip(" \"'“”‘’.,;:!?}>")
        if _looks_like_prepositional_topic_fragment(raw_name):
            return None
        if is_reference_filename_placeholder(raw_name):
            return None
        if looks_like_thematic_lookup_candidate(q, raw_name):
            return None
        if _looks_like_generic_extension_inventory_request(raw_name):
            return None
        if _looks_like_folder_contents_query(raw_name):
            return None
        raw_name_is_identifierish = any(ch in raw_name for ch in ("_", "-", ".", "/", "\\"))
        thematic_doc_request = bool(
            re.search(
                r"(?:关于|有关|相关).*(?:文件|文档|论文|资料)"
                r"|(?:文件|文档|论文|资料).*(?:关于|有关|相关)"
                r"|\b(?:about|related to|regarding|concerning)\b.*\b(files?|documents?|docs?|papers?)\b",
                q,
                re.IGNORECASE,
            )
        )
        if thematic_doc_request and not (has_plausible_filename_extension(raw_name) or raw_name_is_identifierish):
            return None
        if is_descriptive_filename_phrase(raw_name) and not (has_plausible_filename_extension(raw_name) or raw_name_is_identifierish):
            return None
        stem = raw_name
        if has_plausible_filename_extension(raw_name):
            stem = raw_name.rsplit(".", 1)[0].strip()
        file_scope_query = bool(
            re.search(
                r"\b(file|document|doc|image|photo|picture|spreadsheet|table|audio|video)\b",
                q,
                re.IGNORECASE,
            )
            or any(tok in q for tok in ("文件", "文档", "图片", "照片", "表格", "音频", "视频"))
        )
        if len(stem) < 4 and len(raw_name) < 6 and not (
            file_scope_query and (any(ch.isdigit() for ch in raw_name) or any("\u4e00" <= ch <= "\u9fff" for ch in raw_name))
        ) and not (
            extracted_from_filename_surface
            and (
                any("\u4e00" <= ch <= "\u9fff" for ch in raw_name)
                or len(compact_filename_key(stem or raw_name)) >= 4
            )
        ):
            return None

        ql = q.lower()
        is_image = any(k in ql for k in ["image", "photo", "picture", "img", "图片", "照片", "截图"])
        ref = {
            "raw_name": raw_name,
            "search_term": stem or raw_name,
            "kind": "image" if is_image else "file",
        }
        if len(filename_surfaces) >= 2:
            all_names = []
            for surface in filename_surfaces:
                name = os.path.basename(normalize_filename_candidate(str(surface or "").strip()))
                if name and name not in all_names:
                    all_names.append(name)
            if len(all_names) >= 2:
                ref["all_names"] = all_names
        return ref

    def _looks_like_file_content_analysis_query(self, question: str, prompt_language: Optional[str] = None) -> bool:
        from core.intent_analyzer import IntentAnalyzer
        return IntentAnalyzer._looks_like_file_content_analysis_query(question)

    def _extract_file_analysis_focus_query(self, question: str, prompt_language: Optional[str] = None) -> Optional[str]:
        from core.intent_analyzer import IntentAnalyzer
        return IntentAnalyzer._extract_file_analysis_focus_query(question)

    def _normalize_summarize_keyword(
        self,
        category: str,
        raw_keyword: Optional[str],
        original_question: str = "",
    ) -> Optional[str]:
        kw = str(raw_keyword or "").strip()
        if not kw:
            return None

        text = kw
        noise_tokens = [
            str(category or "").strip(),
            "总结", "概括", "归纳", "梳理", "提炼", "汇总",
            "查看", "看", "分析", "讲解", "说明",
            "内容", "主题", "相关",
            "我的", "我", "给我", "帮我", "请", "一下", "下", "并", "进行",
            "文件", "文档", "资料",
            "所有", "全部",
        ]
        for tok in noise_tokens:
            if tok:
                text = text.replace(tok, " ")

        for ch in "，。！？；：、,.!?;:()[]{}<>《》“”\"'`+-_=*/\\|":
            text = text.replace(ch, " ")

        normalized = " ".join(part for part in text.split() if part)
        if not normalized or len(normalized) <= 1:
            return None
        return normalized
    
    def _analyze_intent_with_context(
        self,
        question: str,
        *,
        session_id: Optional[str] = None,
        prompt_language: Optional[str] = None,
        active_paths: Optional[List[str]] = None,
    ) -> dict:
        lang = self._resolve_prompt_language(prompt_language, question=question, session_id=session_id)
        actual_user_lang = self._resolve_prompt_language(None, question=question, session_id=session_id)
        llm = self._get_llm_service(session_id=session_id, prompt_language=lang)
        
        hist_ref = self._get_history_ref(session_id)
        history = []
        _MAX_HIST_RESP_CHARS = 600
        # IMPORTANT: hist_ref[-1] is the CURRENT question just appended by dispatch.
        # We must use hist_ref[-3:-1] to get the true previous 2 turns.
        # Using hist_ref[-2:] would include the current question as "history",
        # causing the Router to mistakenly detect prior context on brand-new chats.
        for h in hist_ref[-3:-1]:
            if h.get('q'):
                history.append({"role": "user", "content": h['q']})
            if h.get('a'):
                a_text = str(h['a'])
                if len(a_text) > _MAX_HIST_RESP_CHARS:
                    a_text = a_text[:_MAX_HIST_RESP_CHARS] + "...[truncated]"
                history.append({"role": "assistant", "content": a_text})
        
        last_results = self._get_last_search_results_ref(session_id)
        category_info = self._get_category_stats(prompt_language=lang)
        
        from core.intent_analyzer import IntentContext, IntentAnalyzer
        ctx = IntentContext(
            question=question,
            prompt_language=lang,
            user_lang=actual_user_lang,
            history=history,
            last_results=last_results,
            get_category_keywords_fn=self._get_rule_category_keywords,
            is_generic_category_fn=self._is_generic_file_scope_category,
            normalize_category_fn=self._normalize_category_name,
            llm_service=llm,
            category_info=category_info,
            prompt_formatter=self._prompt,
            log_followup_guard_fn=self._log_followup_guard,
            session_id=session_id,
            active_paths=active_paths,
        )

        # High-confidence deterministic fast-paths should bypass the LLM router.
        # This keeps selected-scope count/summary and global-summary requests from
        # drifting into search on alternate code paths.
        from core.intent.validator import IntentValidator
        deterministic = IntentValidator.validate(
            question,
            {"action": "search", "params": {"query": question}},
            last_results=last_results,
            history=history,
            active_paths=active_paths,
            prompt_language=lang,
        )
        deterministic_action = str(deterministic.get("action") or "").strip()
        deterministic_scope = str((deterministic.get("params") or {}).get("_scope") or "").strip().lower()
        if deterministic_action == "summarize_all" or (
            deterministic_action == "count" and deterministic_scope == "selected"
        ):
            return self._normalize_intent_to_internal_en(question, deterministic, session_id=session_id)
        
        followup_hint = self._get_followup_hint(session_id)
        if followup_hint:
            from core.intent.followup_hint_guard import should_honor_followup_hint
            if should_honor_followup_hint(
                question,
                followup_hint,
                has_last_results=bool(last_results),
                prompt_language=lang,
            ):
                hinted_action = str(followup_hint.get("action") or "").strip()
                hinted_params = followup_hint.get("params") or {}
                self._consume_followup_hint(session_id)
                return {"action": hinted_action, "params": hinted_params}

        result = IntentAnalyzer.analyze(ctx)
        return self._normalize_intent_to_internal_en(question, result, session_id=session_id)
    
    # _correct_intent has been removed as it is now handled by IntentAnalyzer pipeline.
    
    def _get_category_stats(self, prompt_language: Optional[str] = None) -> str:
        try:
            lang = normalize_prompt_language(prompt_language, fallback="en")
            category_counts = self._collect_category_counts()
            if not category_counts:
                return "- No category stats yet" if lang == "en" else "- 暂无分类统计"

            sorted_cats = sorted(category_counts.items(), key=lambda x: (-int(x[1]), str(x[0])))
            result = ""
            for cat, count in sorted_cats[:30]:
                display_cat = _localize_category_label(str(cat), lang)
                suffix = " files" if lang == "en" else "份"
                result += f"- {display_cat}: {count}{suffix}\n"
            return result.rstrip("\n")
        except Exception as e:
            logger.error(f"获取分类统计失败: {e}")
            return "- other"
    
    def _analyze_intent_with_tools(self, question: str, *, session_id: Optional[str] = None) -> dict:
        results_summary = ""
        session_results = self._get_last_search_results_ref(session_id)
        if session_results:
            results_summary = "Previous search results:\n"
            for i, doc in enumerate(session_results[:10], 1):
                results_summary += f"{i}. {doc.get('file_name', '')} - {doc.get('doc_summary', '')[:50]}\n"
        
        history_text = ""
        hist_ref = self._get_history_ref(session_id)
        if hist_ref:
            for h in hist_ref[-3:]:
                history_text += f"User: {h['q']}\n"
                if h.get('a'):
                    history_text += f"Assistant: {h['a']}\n"
        
        prompt = f"""[Task] Analyze the user's intent and choose one action.

[Conversation History]
{history_text if history_text else "None"}

[Current Input]
{question}

{results_summary}

[Available Actions]
1. search - search documents (params: query=core query, category=optional category)
2. count - count or list selected files
3. view_detail - inspect/summarize one result ("item X", "summarize item X")
4. open_file - open a file only when explicitly requested
5. chat - small talk

[Output Format] JSON only
{{"action": "search", "params": {{"query": "query"}}}}
{{"action": "count", "params": {{"category": "category"}}}}
{{"action": "view_detail", "params": {{"index": 1}}}}
{{"action": "open_file", "params": {{"file_name": "name"}}}}
{{"action": "chat", "params": {{}}}}

        Output JSON only:"""

        try:
            llm = self._get_llm_service(session_id=session_id, prompt_language="en")
            response = llm.generate(prompt)
            
            import json
            import re
            match = re.search(r'\{.*\}', response, re.DOTALL)
            if match:
                result = json.loads(match.group())
                return self._normalize_intent_to_internal_en(question, result)
        except Exception as e:
            logger.error(f"意图分析失败: {e}")

        logger.info("执行意图识别完成---------------")
        
        return self._normalize_intent_to_internal_en(question, {"action": "search", "params": {"query": question}})
    
    def _read_file_content(self, file_path: str) -> str:
        try:
            ext = os.path.splitext(file_path)[1].lower()
            
            if ext == '.pdf':
                if HAS_PDF_TEXT:
                    text, _ = extract_pdf_text(file_path, max_chars=5000)
                    return text[:5000]
            elif ext in ['.txt', '.md']:
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    return f.read()[:5000]
            elif ext == '.docx':
                try:
                    import docx2txt
                    return docx2txt.process(file_path)[:5000]
                except:
                    pass
            
            return "无法读取该文件格式的详细内容"
        except Exception as e:
            return f"读取失败: {e}"
    
    def _open_file(self, file_path: str) -> bool:
        import subprocess
        import platform
        
        try:
            if platform.system() == 'Darwin':  # macOS
                subprocess.run(['open', file_path])
            elif platform.system() == 'Windows':
                os.startfile(file_path)
            else:  # Linux
                subprocess.run(['xdg-open', file_path])
            return True
        except Exception as e:
            logger.error(f"打开文件失败: {e}")
            return False
    
    def _open_file_in_viewer(self, file_path: str, parent_window=None) -> str:
        ext = os.path.splitext(file_path)[1].lower()
        content = ""
        
        try:
            if ext == '.pdf':
                if HAS_PDF_TEXT:
                    try:
                        content, _ = extract_pdf_text(file_path)
                        if not content.strip():
                            content = "[PDF 无法提取文本，建议用系统程序打开]"
                    except Exception as pdf_err:
                        content = f"[PDF 解析失败: {pdf_err}，建议用系统程序打开]"
                else:
                    content = "需要安装 pypdf 才能查看 PDF 文本"
            
            elif ext == '.docx':
                try:
                    import docx2txt
                    content = docx2txt.process(file_path)
                except ImportError:
                    content = "需要安装 docx2txt 才能查看 Word 文档"
            
            elif ext in ['.txt', '.md', '.csv', '.xml']:
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read()
            
            elif ext in ['.xlsx', '.xls']:
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(file_path, data_only=True)
                    for sheet in wb.sheetnames:
                        ws = wb[sheet]
                        content += f"\n=== Sheet: {sheet} ===\n"
                        for row in ws.iter_rows(values_only=True):
                            content += " | ".join([str(cell) if cell else "" for cell in row]) + "\n"
                except ImportError:
                    content = "需要安装 openpyxl 才能查看 Excel"
            
            elif ext in ['.png', '.jpg', '.jpeg', '.gif', '.webp']:
                content = f"[图片文件] {file_path}\n\n暂不支持在文本窗口中显示图片，请使用系统程序打开。"
            
            else:
                try:
                    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                        content = f.read()
                except:
                    content = f"不支持的文件格式: {ext}"
            
            return content if content else "文件内容为空"
            
        except Exception as e:
            return f"读取文件失败: {e}"
    
    def _extract_lexical_features(self, query: str, *, session_id: Optional[str] = None) -> dict:
        if len(query) > 500:
            query = query[:500]

        valid_extensions = {
            ".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".pdf", ".csv", ".tsv",
            ".xlsx", ".xls", ".numbers", ".doc", ".docx", ".ppt", ".pptx", ".md",
            ".txt", ".html", ".json", ".xml", ".wav", ".mp3", ".m4a", ".mp4", ".mov",
        }

        def _merge_unique(values: List[str]) -> List[str]:
            merged: List[str] = []
            seen: set[str] = set()
            for item in values:
                value = str(item or "").strip()
                if not value:
                    continue
                value = _collapse_repeated_surface(value)
                if not value:
                    continue
                key = compact_filename_key(value) or value.lower()
                if key in seen:
                    continue
                seen.add(key)
                merged.append(value)
            return merged

        def _normalize_extensions(values: List[str]) -> List[str]:
            normalized: List[str] = []
            for item in values:
                value = str(item or "").strip().lower()
                if not value:
                    continue
                if not value.startswith("."):
                    value = f".{value}"
                if value in valid_extensions:
                    normalized.append(value)
            return _merge_unique(normalized)

        def _collapse_repeated_surface(text: str) -> str:
            value = " ".join(str(text or "").strip().split())
            if not value:
                return ""
            parts = value.split(" ")
            if len(parts) <= 1:
                return value
            for width in range(1, (len(parts) // 2) + 1):
                if len(parts) % width != 0:
                    continue
                unit = parts[:width]
                if unit and all(parts[idx : idx + width] == unit for idx in range(0, len(parts), width)):
                    return " ".join(unit)
            return value

        def _deterministic_lexical_features(text: str) -> dict:
            import re

            raw = str(text or "").strip()
            if not raw:
                return {"filenames": [], "extensions": []}

            filename_matches: List[str] = []
            extensions: List[str] = []

            def _looks_like_plain_topic_candidate(candidate: str) -> bool:
                value = " ".join(str(candidate or "").strip().split())
                if not value:
                    return False
                if (
                    has_plausible_filename_extension(value)
                    or any(ch.isdigit() for ch in value)
                    or any(ch in value for ch in ("_", "-", "/", "\\", "(", ")", "（", "）"))
                    or any("\u4e00" <= ch <= "\u9fff" for ch in value)
                ):
                    return False
                tokens = [tok for tok in re.findall(r"[a-z]+", value.lower()) if tok]
                if not tokens:
                    return False
                possessives = {"my", "our", "your", "their"}
                if tokens[0] in possessives and len(tokens) <= 4:
                    return True
                plain_topic_terms = {
                    "dog", "dogs", "cat", "cats", "animal", "animals",
                    "photo", "photos", "image", "images", "picture", "pictures",
                    "video", "videos", "song", "songs", "music",
                }
                return len(tokens) == 1 and tokens[0] in plain_topic_terms

            def _looks_like_generic_extension_inventory_request(candidate: str = "") -> bool:
                # Delegates to module-level helper (deduplicated from _extract_explicit_file_reference)
                return _is_generic_ext_inventory_query(raw, candidate)

            for candidate in extract_filelike_candidates(raw, max_candidates=12):
                if not candidate:
                    continue
                if _looks_like_prepositional_topic_fragment(candidate):
                    continue
                lower = candidate.lower()
                if lower.startswith(("http://", "https://", "www.")) or "@" in candidate:
                    continue
                if looks_like_thematic_lookup_candidate(raw, candidate):
                    continue
                if _looks_like_generic_extension_inventory_request(candidate):
                    continue
                stem = os.path.splitext(os.path.basename(candidate))[0]
                if not stem:
                    continue
                normalized_candidate = _collapse_repeated_surface(os.path.basename(candidate))
                if _looks_like_plain_topic_candidate(normalized_candidate):
                    continue
                filename_matches.append(normalized_candidate)

            file_context = re.search(
                r"(file|files|filename|filenames|document|documents|image|images|photo|photos|picture|pictures|spreadsheet|table|tables|audio|video|文件|文档|图片|照片|表格|音频|视频)",
                raw,
                flags=re.IGNORECASE,
            )
            ext_inventory_context = bool(file_context or _is_generic_ext_inventory_query(raw, raw))
            if ext_inventory_context:
                for token in re.findall(r"[a-z0-9]+", raw.lower()):
                    for ext in _INVENTORY_EXTENSION_GROUPS.get(token, ()):
                        if ext:
                            extensions.append(ext)
                cjk_extension_groups = {
                    "表格": (".xlsx", ".xls", ".csv", ".tsv", ".numbers"),
                    "数据表": (".xlsx", ".xls", ".csv", ".tsv", ".numbers"),
                    "电子表格": (".xlsx", ".xls", ".csv", ".tsv", ".numbers"),
                    "工作表": (".xlsx", ".xls", ".csv", ".tsv", ".numbers"),
                }
                for marker, marker_exts in cjk_extension_groups.items():
                    if marker in raw:
                        extensions.extend(marker_exts)

            explicit_ref = self._extract_explicit_file_reference(raw)
            if explicit_ref:
                raw_name = str(explicit_ref.get("raw_name") or "").strip()
                search_term = str(explicit_ref.get("search_term") or "").strip()
                if raw_name and not _looks_like_plain_topic_candidate(raw_name):
                    filename_matches.append(_collapse_repeated_surface(raw_name))
                if (
                    search_term
                    and search_term.lower() != raw_name.lower()
                    and not _looks_like_plain_topic_candidate(search_term)
                ):
                    filename_matches.append(_collapse_repeated_surface(search_term))

            collapsed_raw = _collapse_repeated_surface(raw)
            if (
                collapsed_raw
                and looks_like_specific_filename_candidate(collapsed_raw)
                and not _looks_like_generic_extension_inventory_request(collapsed_raw)
                and not re.search(r"[A-Za-z]", raw)
            ):
                filename_matches.append(collapsed_raw)

            return {
                "filenames": _merge_unique(filename_matches),
                "extensions": _normalize_extensions(extensions),
            }

        deterministic = _deterministic_lexical_features(query)

        # ── Short-circuit: skip LLM when deterministic already found results ──
        # The LLM call costs a full inference pass (~5-20 s on local GGUF).
        # If the regex/heuristic path already identified filenames or extensions
        # (which covers the vast majority of real queries), we can return immediately.
        # Set FILEAGENT_LEXICAL_LLM=1 to force-enable the LLM path if needed.
        _lexical_llm_enabled = str(
            os.getenv("FILEAGENT_LEXICAL_LLM", "0") or "0"
        ).strip().lower() in {"1", "true", "yes", "on"}
        if not _lexical_llm_enabled:
            # LLM path disabled (default): deterministic only
            return deterministic
        if deterministic["filenames"] or deterministic["extensions"]:
            # Deterministic already found something; LLM call would only add marginal
            # value while costing a full inference pass.
            return deterministic

        try:
            lang = self._resolve_prompt_language(None, session_id=session_id)
            prompt = self._prompt("LEXICAL_FEATURE_EXTRACTION_PROMPT", lang).format(query=query)
            llm = self._get_llm_service(detailed=False, session_id=session_id, prompt_language=lang)
            text = llm.generate(prompt).strip()

            import re
            import json

            def _extract_json(s: str):
                if not s: return None
                if (s.startswith("{") and s.endswith("}")): return s
                m = re.search(r"```json\s*([\s\S]*?)\s*```", s, flags=re.IGNORECASE)
                if m: return m.group(1).strip()
                start = s.find("{")
                end = s.rfind("}")
                if start != -1 and end != -1 and end > start:
                    return s[start:end+1]
                return None

            clean_json = _extract_json(text)
            if clean_json:
                parsed = json.loads(clean_json)
                return {
                    "filenames": _merge_unique((parsed.get("filenames", []) or []) + deterministic["filenames"]),
                    "extensions": _normalize_extensions((parsed.get("extensions", []) or []) + deterministic["extensions"]),
                }
        except Exception as e:
            logger.warning(f"Lexical extraction failed: {e}")

        return deterministic
    
    def _analyze_query_intent(self, question: str, *, prompt_language: Optional[str] = None, session_id: Optional[str] = None) -> dict:
        # Keep internal reasoning in English even on legacy branches.
        lang = "en"
        
        hist_ref = self._get_history_ref(session_id)
        history = hist_ref[:-1] if hist_ref else []
        
        if history:
            history_text = " → ".join([h['q'] for h in history[-3:]])
            prompt = self._prompt("REWRITE_QUERY_PROMPT", lang).format(history_context=history_text, current_query=question)
            try:
                llm = self._get_llm_service(session_id=session_id, prompt_language=lang)
                merged = llm.generate(prompt).strip()
                if merged:
                    logger.info(f"LLM整合查询: {merged}")
                    return {"clear": True, "query": merged}
            except Exception as e:
                logger.error(f"LLM整合失败: {e}")
            
            merged = "".join([h['q'] for h in history[-3:]]) + question
            return {"clear": True, "query": merged}
        
        q = question.strip()
        if len(q) <= 4:
            intent_words = [
                "resume", "report", "document", "contract", "employee", "how many", "who", "which",
                "简历", "报告", "文档", "合同", "员工", "有多少", "是谁", "哪些",
            ]
            dynamic_category_words = self._get_rule_category_keywords()
            if not any(w in q for w in intent_words) and not any(c in q for c in dynamic_category_words):
                return {
                    "clear": False,
                    "clarify": (
                        "我还没看明白你具体想找什么。你可以补充关键词、文件类型、文件名，或者说明你想让我总结什么。"
                        if any("\u4e00" <= ch <= "\u9fff" for ch in q)
                        else "I can't tell what you want me to look for yet. You can provide keywords, file type, a file name, or say what you want summarized."
                    ),
                }
        
        return {"clear": True, "query": question}
    
    def _build_context(self, session_id: Optional[str] = None) -> str:
        hist_ref = self._get_history_ref(session_id)
        if not hist_ref:
            return ""
        context = "[Conversation History]\n"
        for h in hist_ref[-5:]:
            context += f"User: {h['q']}\nAssistant: {h['a']}\n"
        return context
    
    def _get_llm_service(
        self,
        detailed: bool = False,
        session_id: Optional[str] = None,
        prompt_language: Optional[str] = None,
    ):
        import os
        resolved_lang = self._resolve_prompt_language(prompt_language, session_id=session_id)
        system_prompt = (
            self._prompt("DETAILED_SYSTEM_PROMPT", resolved_lang)
            if detailed
            else self._prompt("BRIEF_SYSTEM_PROMPT", resolved_lang)
        )

        class _LocalTextService:
            def __init__(self, sys_prompt: str, session_id: Optional[str] = None):

                self._sys = sys_prompt
                self._session_id = session_id
                self._client = get_inproc_openai_client()
                self._last_finish_reason: Optional[str] = None

            def get_last_finish_reason(self) -> Optional[str]:
                fr = str(getattr(self, "_last_finish_reason", "") or "").strip().lower()
                return fr or None

            @staticmethod
            def _is_ctx_overflow_error(err: Exception) -> bool:
                s = str(err or "").lower()
                return (
                    ("context window" in s and ("requested tokens" in s or "request tokens" in s))
                    or ("exceed context window" in s)
                    or ("n_ctx" in s and ("too large" in s or "exceeded" in s))
                    or ("max context length" in s)
                    or ("context shift" in s)
                )

            @staticmethod
            def _smart_shrink_text(text: str, target_len: int) -> str:
                raw = str(text or "")
                if len(raw) <= target_len:
                    return raw
                if target_len <= 64:
                    return raw[:target_len]
                keep_head = int(target_len * 0.65)
                keep_tail = max(0, target_len - keep_head - 20)
                return raw[:keep_head] + "\n...[truncated]...\n" + raw[-keep_tail:]

            def _prompt_budget_tokens(self) -> int:

                try:
                    from services.local_llm import get_local_llm_manager
                    mgr = get_local_llm_manager()
                    n_ctx = int(getattr(mgr, "default_n_ctx", 5120) or 5120)
                except Exception:
                    n_ctx = int(os.getenv("FILEAGENT_LLM_N_CTX", "5120") or 5120)
                reserve = int(os.getenv("FILEAGENT_PROMPT_RESERVE_TOKENS", "640") or 640)
                n_ctx = max(1024, min(n_ctx, 32768))
                reserve = max(128, min(reserve, n_ctx // 2))
                return max(512, n_ctx - reserve)

            def _output_budget_tokens(self, msgs: List[Dict[str, Any]]) -> int:

                try:
                    from services.local_llm import get_local_llm_manager
                    mgr = get_local_llm_manager()
                    n_ctx = int(getattr(mgr, "default_n_ctx", 5120) or 5120)
                except Exception:
                    n_ctx = int(os.getenv("FILEAGENT_LLM_N_CTX", "5120") or 5120)
                n_ctx = max(1024, min(n_ctx, 32768))
                prompt_tokens = self._estimate_prompt_tokens(msgs)
                hard_cap = int(os.getenv("FILEAGENT_MAX_OUTPUT_TOKENS", "1600") or 1600)
                hard_cap = max(96, min(hard_cap, 4096))
                available = max(96, n_ctx - prompt_tokens - 64)
                return max(96, min(hard_cap, available))

            def _estimate_prompt_tokens(self, msgs: List[Dict[str, Any]]) -> int:

                try:
                    chars_per_token = float(os.getenv("LLM_CHARS_PER_TOKEN_EST", "1.6") or 1.6)
                except Exception:
                    chars_per_token = 1.6
                chars_per_token = max(1.1, min(chars_per_token, 4.0))
                total = 0
                for m in (msgs or []):
                    c = str((m or {}).get("content") or "")
                    total += 8
                    total += int(len(c) / chars_per_token)
                total += 24
                return total

            def _shrink_messages_once(self, msgs: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
                out = [dict(m or {}) for m in (msgs or [])]
                if not out:
                    return out

                if len(out) > 2:
                    out.pop(1)
                    return out

                if len(out) >= 1 and out[0].get("role") == "system":
                    c = str(out[0].get("content") or "")
                    if len(c) > 1000:
                        out[0]["content"] = self._smart_shrink_text(c, int(len(c) * 0.82))
                        return out

                for i in range(len(out) - 1, -1, -1):
                    if out[i].get("role") == "user":
                        c = str(out[i].get("content") or "")
                        if len(c) > 1200:
                            out[i]["content"] = self._smart_shrink_text(c, int(len(c) * 0.8))
                            return out
                return out

            def _clip_messages_to_budget(self, msgs: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
                out = [dict(m or {}) for m in (msgs or [])]
                target = self._prompt_budget_tokens()
                guard = 0
                while self._estimate_prompt_tokens(out) > target and guard < 32:
                    nxt = self._shrink_messages_once(out)
                    if nxt == out:
                        break
                    out = nxt
                    guard += 1
                return out

            def _build_messages(self, message: str, history: List[Dict] = None, system_prompt: str = None):

                sys_p = system_prompt or self._sys
                msgs = []
                if sys_p:
                    msgs.append({"role": "system", "content": sys_p})
                if history:
                    max_hist = int(os.getenv("LLM_MAX_HISTORY_TURNS", "8"))
                    recent = (history or [])[-max_hist:]
                    for h in recent:
                        try:
                            r = (h or {}).get("role")
                            c = (h or {}).get("content")
                        except Exception:
                            r, c = None, None
                        if r not in {"user", "assistant"}:
                            continue
                        if not c or not isinstance(c, str) or not str(c).strip():
                            continue
                        content_str = str(c).strip()
                        if len(content_str) == 0:
                            continue
                        max_single = int(os.getenv("LLM_MAX_SINGLE_MSG_CHARS", "4000"))
                        if len(content_str) > max_single:
                            content_str = content_str[:max_single] + "..."
                        msgs.append({"role": r, "content": content_str})
                msgs.append({"role": "user", "content": message})
                return self._clip_messages_to_budget(msgs)

            def generate(
                self,
                message: str,
                history: List[Dict] = None,
                system_prompt: str = None,
                temperature: Optional[float] = None,
                max_tokens: Optional[int] = None,
                stop: Optional[List[str]] = None,
                **_: Any,
            ) -> str:

                sys_p = system_prompt or self._sys
                msgs = self._build_messages(message, history=history, system_prompt=sys_p)
                max_retry = int(os.getenv("FILEAGENT_CTX_RETRY_MAX", "3") or 3)
                sync_token_cap = int(os.getenv("FILEAGENT_SYNC_GENERATE_MAX_TOKENS", "256") or 256)
                sync_stop_markers = [
                    "<think>",
                    "</think>",
                    "<|channel>thought",
                    "<channel|>thought",
                    "<|channel|>analysis",
                    "<channel|>analysis",
                ]
                merged_stop = list(dict.fromkeys(list(stop or []) + sync_stop_markers))
                for attempt in range(max_retry + 1):
                    abort_mgr = get_global_abort_manager()
                    if abort_mgr.is_aborted(self._session_id):
                        return ""
                    try:
                        output_budget = self._output_budget_tokens(msgs)
                        if max_tokens is not None:
                            try:
                                output_budget = min(output_budget, int(max_tokens))
                            except Exception:
                                pass
                        if sync_token_cap > 0:
                            output_budget = min(output_budget, sync_token_cap)
                        it = self._client.chat.completions.create(
                            model=None,
                            messages=msgs,
                            stream=True,
                            temperature=float(0.2 if temperature is None else temperature),
                            max_tokens=output_budget,
                            stop=merged_stop,
                        )
                        collected = []
                        for ev in it:
                            if abort_mgr.is_aborted(self._session_id):
                                logger.info(f"检测到中断标志，停止同步生成 (session={self._session_id})")
                                try:
                                    it.close()
                                except Exception:
                                    pass
                                return "".join(collected).strip()
                            try:
                                delta = ((ev.choices or [{}])[0].delta.content) or ""
                                if delta:
                                    collected.append(delta)
                            except Exception:
                                pass
                        return "".join(collected).strip()
                    except Exception as e:
                        if self._is_ctx_overflow_error(e) and attempt < max_retry:
                            logger.warning(
                                f"[LLM] context overflow (sync), shrinking prompt and retrying "
                                f"(attempt={attempt + 1}/{max_retry})"
                            )
                            nxt = self._shrink_messages_once(msgs)
                            if nxt == msgs:
                                break
                            msgs = nxt
                            continue
                        raise
                return ""

            def generate_stream(self, message: str, history: List[Dict] = None, system_prompt: str = None):

                sys_p = system_prompt or self._sys
                msgs = self._build_messages(message, history=history, system_prompt=sys_p)
                max_retry = int(os.getenv("FILEAGENT_CTX_RETRY_MAX", "3") or 3)
                attempt = 0
                self._last_finish_reason = None
                while attempt <= max_retry:
                    it = None
                    emitted_any = False
                    _in_think_block = False
                    _think_buf = ""
                    try:
                        it = self._client.chat.completions.create(
                            model=None,
                            messages=msgs,
                            stream=True,
                            temperature=0.2,
                            max_tokens=self._output_budget_tokens(msgs),
                        )
                        abort_mgr = get_global_abort_manager()
                        for ev in it:
                            if abort_mgr.is_aborted(self._session_id):
                                logger.info(f"检测到中断标志，停止流式生成 (session={self._session_id})")
                                try:
                                    it.close()
                                except Exception:
                                    pass
                                return
                            try:
                                fr = getattr(ev.choices[0], "finish_reason", None)
                                if fr is not None:
                                    self._last_finish_reason = str(fr).strip().lower()
                            except Exception:
                                pass
                            try:
                                delta = ev.choices[0].delta.content or ""
                            except Exception:
                                delta = ""
                            if delta:
                                _think_buf += delta
                                while _think_buf:
                                    if _in_think_block:
                                        end_idx = _think_buf.find("</think>")
                                        if end_idx >= 0:
                                            _think_buf = _think_buf[end_idx + 8:]
                                            _in_think_block = False
                                            continue
                                        else:
                                            _think_buf = ""
                                            break
                                    else:
                                        start_idx = _think_buf.find("<think>")
                                        if start_idx >= 0:
                                            before = _think_buf[:start_idx]
                                            if before:
                                                emitted_any = True
                                                yield before
                                            _think_buf = _think_buf[start_idx + 7:]
                                            _in_think_block = True
                                            continue
                                        else:
                                            safe_len = max(0, len(_think_buf) - 7)
                                            if safe_len > 0:
                                                emitted_any = True
                                                yield _think_buf[:safe_len]
                                            _think_buf = _think_buf[safe_len:]
                                            break
                        if _think_buf and not _in_think_block:
                            emitted_any = True
                            yield _think_buf
                        return
                    except Exception as e:
                        if (not emitted_any) and self._is_ctx_overflow_error(e) and attempt < max_retry:
                            logger.warning(
                                f"[LLM] context overflow (stream), shrinking prompt and retrying "
                                f"(attempt={attempt + 1}/{max_retry})"
                            )
                            nxt = self._shrink_messages_once(msgs)
                            if nxt == msgs:
                                break
                            msgs = nxt
                            attempt += 1
                            continue
                        self._last_finish_reason = "error"
                        text = self.generate(message, history=history, system_prompt=system_prompt)
                        chunk_size = int(os.getenv("LOCAL_LLM_STREAM_CHUNK_CHARS", "120"))
                        for i in range(0, len(text), chunk_size):
                            yield text[i : i + chunk_size]
                        return
                    finally:
                        if it is not None:
                            try:
                                it.close()
                            except Exception:
                                pass

        # Do not reuse a single cached service across sessions/languages.
        # session_id is used for abort control, and system prompt depends on language.
        try:
            from services.local_llm import get_local_llm_manager
            _llm_mgr = self._llm_manager or get_local_llm_manager()
            _mid = getattr(_llm_mgr, "current_model_id", None)
            if _mid:
                _cfg = _llm_mgr.get_target_model_config(_mid) or {}
                _suffix = str(_cfg.get("intent_prompt_suffix") or "").strip()
                if _suffix and not system_prompt.rstrip().endswith(_suffix):
                    system_prompt = f"{system_prompt}\n{_suffix}"
        except Exception:
            pass
        return _LocalTextService(system_prompt, session_id=session_id)
    
    def _generate_clarify_question(self, question: str, *, prompt_language: Optional[str] = None, session_id: Optional[str] = None) -> str:
        lang = self._resolve_prompt_language(prompt_language, question=question, session_id=session_id)
        prompt = self._prompt("AMBIGUOUS_QUERY_PROMPT", lang).format(question=question)
        
        llm = self._get_llm_service(prompt_language=lang, session_id=session_id)
        return llm.generate(prompt)
    
    def _find_file_by_name_or_index(self, index: int = 0, file_name: str = "", session_id: Optional[str] = None) -> dict:
        last_results = self._get_last_search_results_ref(session_id)
        if not last_results:
            return None
        
        if index > 0 and index <= len(last_results):
            return last_results[index - 1]
        
        if file_name:
            for doc in last_results:
                if file_name.lower() in doc.get('file_name', '').lower():
                    return doc
                if file_name.lower() in doc.get('doc_summary', '').lower():
                    return doc
        
        return None
    
    def _handle_view_detail(
        self,
        index: int,
        file_name: str,
        session_id: Optional[str] = None,
        prompt_language: Optional[str] = None,
    ):
        """Handle view-detail request — delegates to view_detail_handler (media-aware)."""
        from core.handlers.view_detail_handler import _handle_view_detail as _dh
        yield from _dh(self, index, file_name,
                       session_id=session_id, prompt_language=prompt_language)


    def _extract_ext_filters(self, question: str, params: dict = None) -> List[str]:
        """Extract file extension filters from query text (also merges params[file_extensions]).
    
        _extract_ext_filters_simple delegates here — single source of truth.
        """
        import re
        q_lower = (question or "").lower()
        tokens = set(re.findall(r'[a-z0-9]+', q_lower))
        exts: List[str] = []
        if "csv" in tokens: exts.append(".csv")
        if "pdf" in tokens: exts.append(".pdf")
        if "txt" in tokens or "text" in tokens: exts.append(".txt")
        if "excel" in tokens or "xlsx" in tokens or "xls" in tokens: exts.extend([".xlsx", ".xls"])
        if "word" in tokens or "docx" in tokens: exts.extend([".docx", ".doc"])
        elif "doc" in tokens: exts.append(".doc")
        if "ppt" in tokens or "pptx" in tokens or "powerpoint" in tokens: exts.extend([".pptx", ".ppt"])
        if "md" in tokens or "markdown" in tokens: exts.append(".md")
        if "jpg" in tokens or "jpeg" in tokens: exts.extend([".jpg", ".jpeg"])
        if "png" in tokens: exts.append(".png")
        if "mp3" in tokens: exts.append(".mp3")
        if "mp4" in tokens: exts.append(".mp4")
        if params and params.get("file_extensions"):
            for e in params.get("file_extensions").split(","):
                e = e.strip().lower()
                if not e: continue
                if not e.startswith("."): e = "." + e
                if e not in exts: exts.append(e)
        return exts

    # ──────────────────────────────────────────────────────────────────────
    # media_timequery handler
    # ──────────────────────────────────────────────────────────────────────

    def _handle_media_export(
        self,
        question: str,
        params: dict,
        *,
        session_id: Optional[str] = None,
        prompt_language: Optional[str] = None,
        active_paths: Optional[List[str]] = None,
    ):
        from core.handlers.media_handler import _handle_media_export as delegated_handler

        yield from delegated_handler(
            self,
            question,
            params,
            session_id=session_id,
            prompt_language=prompt_language,
            active_paths=active_paths,
        )

    # ──────────────────────────────────────────────────────────────────────
    # media_content_search handler
    # ──────────────────────────────────────────────────────────────────────

    def _handle_media_content_search(
        self,
        question: str,
        params: dict,
        *,
        session_id: Optional[str] = None,
        prompt_language: Optional[str] = None,
        active_paths: Optional[List[str]] = None,
    ):
        from core.handlers.media_handler import _handle_media_content_search as delegated_handler

        yield from delegated_handler(
            self,
            question,
            params,
            session_id=session_id,
            prompt_language=prompt_language,
            active_paths=active_paths,
        )

    def _handle_count(self, category: str, original_question: str, allowed_paths: Optional[List[str]] = None, session_id: Optional[str] = None, params: dict = None):
        from core.handlers.count_handler import _handle_count as delegated_handler

        yield from delegated_handler(
            self,
            category,
            original_question,
            allowed_paths=allowed_paths,
            session_id=session_id,
            params=params,
        )
    
    def _handle_summarize(
        self,
        category: str,
        original_question: str,
        *,
        keyword: Optional[str] = None,
        allowed_paths: Optional[List[str]] = None,
        session_id: Optional[str] = None,
        prompt_language: Optional[str] = None,
        params: dict = None,
    ):
        from core.handlers.summarize_handler import _handle_summarize as delegated_handler

        yield from delegated_handler(
            self,
            category,
            original_question,
            keyword=keyword,
            allowed_paths=allowed_paths,
            session_id=session_id,
            prompt_language=prompt_language,
            params=params,
        )

    def _handle_process_previous(
        self,
        original_question: str,
        session_id: Optional[str] = None,
        prompt_language: Optional[str] = None,
        active_paths: Optional[List[str]] = None,
        params: Optional[Dict[str, Any]] = None,
    ):
        from core.handlers.process_previous_handler import _handle_process_previous as delegated_handler

        yield from delegated_handler(
            self,
            original_question,
            session_id=session_id,
            prompt_language=prompt_language,
            active_paths=active_paths,
            params=params,
        )

    def query_stream(
        self,
        question: str,
        session_id: Optional[str] = None,
        prompt_language: Optional[str] = None,
    ):
        logger.info(f"\n[Agent] 流式查询: {question}")
        logger.info(f"mode=intent_dispatch (query_stream)")
        self._log(f"用户输入: {question}", "USER")

        yield from self._query_stream_intent_dispatch(
            question,
            active_paths=None,
            session_id=session_id,
            emit_status=False,
            prompt_language=prompt_language,
        )
        return

    def _extract_ext_filters_simple(self, question: str) -> List[str]:
        """Thin alias for _extract_ext_filters (kept for backward compatibility)."""
        return self._extract_ext_filters(question, params=None)

    def _handle_summarize_all(
        self,
        question: str,
        params: dict,
        active_paths: Optional[List[str]],
        session_id: Optional[str],
        emit_status: bool,
        prompt_language: Optional[str] = None,
    ):
        from core.handlers.summarize_all_handler import _handle_summarize_all as delegated_handler

        yield from delegated_handler(
            self,
            question,
            params,
            active_paths=active_paths,
            session_id=session_id,
            emit_status=emit_status,
            prompt_language=prompt_language,
        )

    def _query_stream_intent_dispatch(
        self,
        question: str,
        *,
        active_paths: Optional[List[str]],
        session_id: Optional[str],
        emit_status: bool,
        prompt_language: Optional[str] = None,
        opened_file_path: Optional[str] = None,
    ):
        from core.agent.dispatch import _query_stream_intent_dispatch as delegated_dispatch

        yield from delegated_dispatch(
            self,
            question,
            active_paths=active_paths,
            session_id=session_id,
            emit_status=emit_status,
            prompt_language=prompt_language,
            opened_file_path=opened_file_path,
        )

    def query_stream_live(
        self,
        question: str,
        active_source_ids: Optional[List[str]] = None,
        active_paths: Optional[List[str]] = None,
        model_id: Optional[str] = None,
        files_preview_k: int = 5,
        session_id: Optional[str] = None,
        prompt_language: Optional[str] = None,
    ):
        from core.agent.live_stream import query_stream_live as delegated_live_stream

        yield from delegated_live_stream(
            self,
            question,
            active_source_ids=active_source_ids,
            active_paths=active_paths,
            model_id=model_id,
            files_preview_k=files_preview_k,
            session_id=session_id,
            prompt_language=prompt_language,
        )
    
    def _collect_sources(self, question: str, category_filter: str = None) -> List[Dict]:
        kb = get_kb_instance()
        results = kb.vector_search(question, n_results=settings.VECTOR_SEARCH_TOP_K)
        reranked = kb.rerank(question, results, top_k=settings.RERANK_TOP_K)
        
        file_best = {}  # {file_path: doc}
        for doc in reranked:
            if doc.get('rerank_score', 0) >= settings.RELEVANCE_THRESHOLD:
                file_path = doc.get('file_path', '')
                if file_path not in file_best or doc.get('rerank_score', 0) > file_best[file_path].get('rerank_score', 0):
                    file_best[file_path] = {
                        'file_name': doc.get('file_name', ''),
                        'file_path': file_path,
                        'doc_summary': doc.get('doc_summary', ''),
                        'doc_category': doc.get('doc_category', ''),
                        'rerank_score': doc.get('rerank_score', 0),
                        'text': doc.get('text', ''),
                    }
        
        sources = sorted(file_best.values(), key=lambda x: x.get('rerank_score', 0), reverse=True)
        return sources

    def _handle_tools(self, task: str, session_id: Optional[str] = None):
        from core.agent.tool_agent import _handle_tools as delegated_handle_tools

        yield from delegated_handle_tools(self, task, session_id=session_id)

    def _run_tool_agent(self, task: str, require_tools: bool = False, session_id: Optional[str] = None):
        from core.agent.tool_agent import _run_tool_agent as delegated_run_tool_agent

        yield from delegated_run_tool_agent(self, task, require_tools=require_tools, session_id=session_id)
    
    def _build_graph(self):
        from core.agent.tool_agent import _build_graph as delegated_build_graph

        return delegated_build_graph(self)
    
    def query(self, question: str, session_id: Optional[str] = None) -> Dict[str, Any]:
        answer_parts: List[str] = []
        sources: List[Dict[str, Any]] = []
        trace: List[Dict[str, Any]] = []
        query_type = "search"
        need_clarify = False

        for ev in self.query_stream(question, session_id=session_id):
            ev_type = ev.get("type")
            if ev_type == "text":
                answer_parts.append(ev.get("content", "") or ev.get("delta", ""))
            elif ev_type == "sources":
                sources = ev.get("content", []) or []
            elif ev_type == "trace":
                trace = ev.get("content", []) or []
            elif ev_type == "done":
                query_type = ev.get("query_type", query_type)
                need_clarify = query_type == "clarify"

        return {
            "answer": "".join(answer_parts).strip(),
            "sources": sources,
            "trace": trace,
            "query_type": query_type,
            "need_clarify": need_clarify,
        }
    
    def _query_with_tools(self, question: str) -> Dict[str, Any]:
        logger.info(f"[Agent] 执行统计查询")
        
        kb = get_kb_instance()
        
        all_stats = kb.count_all_categories()
        available_categories = list(all_stats.keys())
        
        detected_category = self._detect_category_with_llm(question, available_categories)
        logger.info(f"[Agent] LLM 识别分类: {detected_category}")
        
        try:
            if detected_category and detected_category != "all":
                result = kb.count_by_category(category=detected_category)
                count = result.get('count', 0)
                answer = f"您有 {count} 份{detected_category}文件。"
            else:
                total = sum(all_stats.values())
                
                stats_lines = [f"- {cat}: {cnt} 份" for cat, cnt in sorted(all_stats.items(), key=lambda x: x[1], reverse=True)]
                answer = f"您共有 {total} 个文件：\n" + "\n".join(stats_lines[:8])
            
            logger.info(f"[Agent] 统计结果: {answer}")
            return {
                "answer": answer,
                "sources": [],
                "query_type": "stats",
            }
        except Exception as e:
            logger.error(f"[Agent] 统计失败: {e}")
            return {
                "answer": f"统计失败: {e}",
                "sources": [],
                "query_type": "stats",
            }
    
    def _detect_category_with_llm(self, question: str, available_categories: list) -> str:
        try:
            from config.prompts import DETECT_CATEGORY_PROMPT
            client = self._get_local_llm_client()
            
            categories_str = "、".join(available_categories)
            
            prompt = DETECT_CATEGORY_PROMPT.format(question=question, categories_str=categories_str)

            response = client.chat.completions.create(
                model=None,   # auto = index/chat model, no VL switching needed
                messages=[{"role": "user", "content": prompt}],
                max_tokens=20,
                temperature=0
            )
            
            result = response.choices[0].message.content.strip()
            
            if result in available_categories:
                return result
            elif "全部" in result or "所有" in result or "all" in result.lower():
                return "all"
            else:
                for cat in available_categories:
                    if cat in result or result in cat:
                        return cat
                return "all"
                
        except Exception as e:
            logger.error(f"[Agent] LLM 分类识别失败: {e}")
            return "all"
    
    
    def _query_with_forced_search(self, question: str, *, session_id: Optional[str] = None) -> Dict[str, Any]:
        
        logger.info(f"Step 1: 执行向量检索")
        search_result = search_documents.invoke({"query": question, "top_k": 10})
        
        if search_result == "[NO_RELEVANT_DOCS]":
            logger.info(f"检索无结果，进入 chat 模式")
            
            llm = get_llm()
            context = self._build_context(session_id)
            sys_prompt = (
                "You are an efficient assistant. Keep the answer concise and focused.\n"
                f"{context}"
            )
            chat_messages = build_messages_for_model(sys_prompt, [], question)
            chat_response = llm.invoke(chat_messages)
            
            return {
                "answer": chat_response.content,
                "sources": [],
                "query_type": "chat",
            }
        
        logger.info(f"检索有结果，生成回答")
        
        sources = []
        kb = get_kb_instance()
        search_results = kb.vector_search(question, n_results=settings.VECTOR_SEARCH_TOP_K)
        reranked = kb.rerank(question, search_results, top_k=settings.RERANK_TOP_K)
        
        for doc in reranked:
            if doc.get('rerank_score', 0) >= settings.RELEVANCE_THRESHOLD:
                sources.append({
                    'file_name': doc.get('file_name', ''),
                    'file_path': doc.get('file_path', ''),
                    'doc_summary': doc.get('doc_summary', ''),
                    'doc_category': doc.get('doc_category', ''),
                    'rerank_score': doc.get('rerank_score', 0),
                    'text': doc.get('text', '')[:300],
                })
        
        if not sources:
            return {
                "answer": "No relevant information was found in indexed content. Please provide a more specific file name or keyword.",
                "sources": [],
                "query_type": "search",
            }
        
        llm = get_llm()
        
        context = self._build_context(session_id)
        
        final_prompt = f"""You are an efficient file assistant. Requirements:
1. Keep the answer concise and accurate.
2. For complex questions, use short bullet points.
3. End with a clear source hint using file names.

{context}
<Retrieved Results>
{search_result}
</Retrieved Results>

<Question>
{question}
</Question>"""

        response = llm.invoke([HumanMessage(content=final_prompt)])
        
        return {
            "answer": response.content,
            "sources": sources,
            "query_type": "search",
        }



if __name__ == "__main__":
    agent = FileAgent()
    
    test_questions = [
        "有多少份简历",
        "报告有哪些",
        "合同有几份",
        "设计总监是谁",
        "帮我找项目相关的文档",
        "你好",
    ]
    
    for q in test_questions:
        logger.info(f"\n{'='*60}")
        result = agent.query(q)
        logger.info(f"\n问题: {q}")
        logger.info(f"类型: {result['query_type']}")
        answer = result['answer']
        logger.info(f"回答: {answer[:300] if len(answer) > 300 else answer}")
        logger.info(f"来源文件: {len(result['sources'])} 个")
